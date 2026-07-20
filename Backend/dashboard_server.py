import io
import os
import sys
import logging
import datetime
import requests
import json
import pandas as pd
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from app.config import settings

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)

# Constants for ERP Connection (credenciales via Backend/.env, ver app/config.py)
BASE_URL = settings.MBA3_BASE_URL_PROD
URL_LOGIN = f"{BASE_URL}/ws2_mba3_serv_/login_servicio"
URL_CONSULTA = f"{BASE_URL}/ws2_mba3_serv_Consultas_Externas_/"

CODIGO_SERVICIO = settings.MBA3_CODIGO_SERVICIO_PROD
PASSWORD_SERVICIO = settings.MBA3_PASSWORD_SERVICIO_PROD

# ----------------- HELPER FUNCTIONS -----------------

def procesar_respuesta_erp(datos, context=""):
    """
    Valida y limpia la respuesta del ERP. Retorna una lista de registros
    válidos o una lista vacía si hay un error o no hay registros.
    """
    if not datos:
        return []
        
    # Si es un diccionario (registro único o error)
    if isinstance(datos, dict):
        if datos.get("codigo") == "009": # Código de "No se encontraron registros"
            logging.info(f"Backend: No se encontraron registros en {context}.")
            return []
        elif "error" in datos or "codigo" in datos:
            logging.error(f"Backend: Error devuelto por ERP en {context}: {datos}")
            return []
        return [datos]
        
    # Si es una lista
    if isinstance(datos, list):
        if len(datos) == 0:
            return []
        # Verificar si el primer elemento es un error/código especial
        first = datos[0]
        if isinstance(first, dict):
            if first.get("codigo") == "009":
                logging.info(f"Backend: No se encontraron registros en {context}.")
                return []
            elif "error" in first or "codigo" in first:
                logging.error(f"Backend: Error devuelto por ERP en lista para {context}: {first}")
                return []
        return datos
        
    return []

def obtener_token():
    """
    Authenticates against MBA3 and returns the JWT token.
    """
    logging.info("Backend: Autenticando con MBA3...")
    headers = {"Content-Type": "application/json"}
    payload = {
        "codigo": CODIGO_SERVICIO,
        "pwd": PASSWORD_SERVICIO
    }
    try:
        response = requests.post(URL_LOGIN, json=payload, headers=headers, timeout=15)
        response.raise_for_status()
        datos = response.json()
        token = datos.get("jwt")
        if token:
            logging.info("Backend: Autenticación exitosa.")
            return token
        else:
            logging.error("Backend: La respuesta no contiene JWT.")
            return None
    except Exception as e:
        logging.error(f"Backend: Error en login: {e}")
        return None

def ejecutar_consulta_tabla(token, columnas, tabla, condicion_where=None, limite=None):
    """
    Executes an ERP query and returns the JSON result.
    """
    headers = {"Authorization": token}
    payload = {
        "select": columnas,
        "from": tabla
    }
    if condicion_where:
        payload["where"] = condicion_where
    if limite:
        payload["limit"] = str(limite)
        
    try:
        # High timeout for massive tables
        response = requests.post(URL_CONSULTA, headers=headers, data=payload, timeout=120)
        response.raise_for_status()
        datos = response.json()
        return procesar_respuesta_erp(datos, f"tabla {tabla}")
    except Exception as e:
        logging.error(f"Backend: Error consultando {tabla}: {e}")
        return []

# ----------------- DATA EXTRACTION FUNCTIONS -----------------

def get_movimientos_df(token, fecha_inicio, fecha_fin):
    logging.info(f"Backend: Consultando movimientos desde {fecha_inicio} hasta {fecha_fin}")
    condicion_where = f"TRANS_DATE >= '{fecha_inicio}' AND TRANS_DATE <= '{fecha_fin}'"
    columnas = "TRANS_DATE,PRODUCT_NAME,Codigo_producto_convertido,ORIGINAL_QTY,ORIGIN_MEMO,ORIGIN_REF,BASE_COMISION,Info_Seriales,Codigo_Sucursal,BaseImponibleReal_1,COD_SALESMAN"
    
    headers = {"Authorization": token}
    payload = {
        "select": columnas,
        "from": "INVT_Producto_Movimientos",
        "where": condicion_where
    }
    
    try:
        response = requests.post(URL_CONSULTA, headers=headers, data=payload, timeout=60)
        response.raise_for_status()
        datos = response.json()
        
        registros = procesar_respuesta_erp(datos, "INVT_Producto_Movimientos")
        df = pd.DataFrame(registros)
        if not df.empty:
            df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
            # Ensure columns exist before converting to numeric
            for col in ['ORIGINAL_QTY', 'BASE_COMISION', 'BaseImponibleReal_1']:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                else:
                    df[col] = 0.0
        return df
    except Exception as e:
        logging.error(f"Backend: Error en consulta de movimientos: {e}")
        return pd.DataFrame()

def get_liquidaciones_df(token, fecha_inicio, fecha_fin):
    logging.info(f"Backend: Consultando liquidaciones desde {fecha_inicio} hasta {fecha_fin}")
    cols_principal = "CORP,LIQUIDACION_FECHA,OBSERVACIONES,ANTES_TOTAL_1,ANTES_TOTAL_2,ANTES_TOTAL_3,DESPUES_TOTAL_1,DESPUES_TOTAL_2,DESPUES_TOTAL_3,LIQUIDACION_ESTADO,LIQUIDACION_ID_CORP"
    where_principal = f"CORP = 'NVC01' AND LIQUIDACION_FECHA >= '{fecha_inicio}' AND LIQUIDACION_FECHA <= '{fecha_fin}'"
    
    datos_principal = ejecutar_consulta_tabla(token, cols_principal, "PROV_Liquidaciones_Principal", where_principal, 5000)
    if not datos_principal:
        return pd.DataFrame()
        
    df_principal = pd.DataFrame(datos_principal)
    
    cols_productos = "FACTURA_ID_CORP,IdRecepcionRelacionada,VALOR_TOTAL_CIF,VALOR_SUBTOTAL_CIF,VALOR_ANTES_1,VALOR_ANTES_2,VALOR_ANTES_3,VALOR_DESPUES_1,VALOR_DESPUES_2,VALOR_DESPUES_3,PARTIDA_ID_CORP,PRODUCTO_ID_CORP,LIQUIDACION_ID,CANTIDAD,PRECIO,TOTAL,VALOR_TOTAL_CIF_MANUAL,VALOR_TOTAL_CIF_UNIDAD,LIQUIDACION_ID_CORP"
    datos_productos = ejecutar_consulta_tabla(token, cols_productos, "PROV_Liquidaciones_Productos", limite=50000)
    if not datos_productos:
        return pd.DataFrame()
        
    df_productos = pd.DataFrame(datos_productos)
    
    # Check if we have join keys
    if 'LIQUIDACION_ID_CORP' not in df_principal.columns or 'LIQUIDACION_ID_CORP' not in df_productos.columns:
        return pd.DataFrame()
        
    # Inner join in RAM
    df_cruce = pd.merge(df_principal, df_productos, on='LIQUIDACION_ID_CORP', how='inner')
    
    columnas_finales = [
        "CORP", "LIQUIDACION_FECHA", "OBSERVACIONES", "FACTURA_ID_CORP",
        "IdRecepcionRelacionada", "ANTES_TOTAL_1", "ANTES_TOTAL_2", "ANTES_TOTAL_3",
        "DESPUES_TOTAL_1", "DESPUES_TOTAL_2", "DESPUES_TOTAL_3", "VALOR_TOTAL_CIF",
        "VALOR_SUBTOTAL_CIF", "VALOR_ANTES_1", "VALOR_ANTES_2", "VALOR_ANTES_3",
        "VALOR_DESPUES_1", "VALOR_DESPUES_2", "VALOR_DESPUES_3", "PARTIDA_ID_CORP",
        "PRODUCTO_ID_CORP", "LIQUIDACION_ID", "CANTIDAD", "PRECIO", "TOTAL",
        "VALOR_TOTAL_CIF_MANUAL", "VALOR_TOTAL_CIF_UNIDAD", "LIQUIDACION_ESTADO",
        "LIQUIDACION_ID_CORP"
    ]
    
    columnas_disponibles = [col for col in columnas_finales if col in df_cruce.columns]
    df_final = df_cruce[columnas_disponibles]
    
    if not df_final.empty:
        df_final = df_final.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        # Parse numbers
        numeric_cols = [
            "ANTES_TOTAL_1", "ANTES_TOTAL_2", "ANTES_TOTAL_3", 
            "DESPUES_TOTAL_1", "DESPUES_TOTAL_2", "DESPUES_TOTAL_3", 
            "VALOR_TOTAL_CIF", "VALOR_SUBTOTAL_CIF", "VALOR_ANTES_1", 
            "VALOR_ANTES_2", "VALOR_ANTES_3", "VALOR_DESPUES_1", 
            "VALOR_DESPUES_2", "VALOR_DESPUES_3", "CANTIDAD", 
            "PRECIO", "TOTAL", "VALOR_TOTAL_CIF_MANUAL", "VALOR_TOTAL_CIF_UNIDAD"
        ]
        for col in numeric_cols:
            if col in df_final.columns:
                df_final[col] = pd.to_numeric(df_final[col], errors='coerce').fillna(0)
            else:
                df_final[col] = 0.0
    return df_final

def get_ats_df(token, fecha_inicio, fecha_fin):
    logging.info(f"Backend: Consultando ATS desde {fecha_inicio} hasta {fecha_fin}")
    
    cols_factura = "INVOICE_DATE,CORP,VENDOR_ID,MEMO,INVOICE_TOTAL,DOC_REFERENCE,TotalProductosConIVa,TotalServiciosConIVa,TotalProductosSinIVa,TotalServiciosSinIVa,VOID,DOC_ID_CORP,CONFIRMED"
    where_factura = f"INVOICE_DATE >= '{fecha_inicio}' AND INVOICE_DATE <= '{fecha_fin}'"
    
    datos_factura = ejecutar_consulta_tabla(token, cols_factura, "PROV_Factura_Principal", where_factura, 100000)
    if not datos_factura:
        return pd.DataFrame()
    df_facturas = pd.DataFrame(datos_factura)
    
    cols_proveedor = "VENDOR_ID,VENDOR_NAME,RUC_or_FED_ID"
    datos_proveedor = ejecutar_consulta_tabla(token, cols_proveedor, "PROV_Ficha_Principal", limite=100000)
    if not datos_proveedor:
        return pd.DataFrame()
    df_proveedores = pd.DataFrame(datos_proveedor)
    
    cols_fiscal = "MF_Nume1,MF_Alfa2,MF_Lista2,MF_Bool5,ID_Relacionado"
    datos_fiscal = ejecutar_consulta_tabla(token, cols_fiscal, "CONT_Info_Fiscal", limite=100000)
    if not datos_fiscal:
        return pd.DataFrame()
    df_fiscal = pd.DataFrame(datos_fiscal)
    
    # Check join columns exist
    if 'VENDOR_ID' not in df_facturas.columns or 'VENDOR_ID' not in df_proveedores.columns:
        return pd.DataFrame()
    if 'DOC_ID_CORP' not in df_facturas.columns or 'ID_Relacionado' not in df_fiscal.columns:
        return pd.DataFrame()
        
    # Inner joins in RAM
    df_cruce = pd.merge(df_facturas, df_proveedores, on='VENDOR_ID', how='inner')
    df_cruce = pd.merge(df_cruce, df_fiscal, left_on='DOC_ID_CORP', right_on='ID_Relacionado', how='inner')
    
    def parse_bool(val):
        if pd.isna(val): return False
        return str(val).strip().lower() in ['true', '1', 't', 'y', 'yes']
        
    if 'CONFIRMED' in df_cruce.columns:
        df_cruce['CONFIRMED_BOOL'] = df_cruce['CONFIRMED'].apply(parse_bool)
    else:
        df_cruce['CONFIRMED_BOOL'] = False
        
    if 'VOID' in df_cruce.columns:
        df_cruce['VOID_BOOL'] = df_cruce['VOID'].apply(parse_bool)
    else:
        df_cruce['VOID_BOOL'] = False
        
    if 'MF_Bool5' in df_cruce.columns:
        df_cruce['MF_Bool5_BOOL'] = df_cruce['MF_Bool5'].apply(parse_bool)
    else:
        df_cruce['MF_Bool5_BOOL'] = False
    
    df_cruce = df_cruce[
        (df_cruce['CONFIRMED_BOOL'] == True) &
        (df_cruce['VOID_BOOL'] == False) &
        (df_cruce['MF_Bool5_BOOL'] == False)
    ]
    
    df_cruce['ES_ANULADO'] = df_cruce['VOID_BOOL'].apply(lambda x: 1 if x else 0)
    
    campos_numericos = ['TotalProductosConIVa', 'TotalServiciosConIVa', 'TotalProductosSinIVa', 'TotalServiciosSinIVa', 'INVOICE_TOTAL']
    for col in campos_numericos:
        if col in df_cruce.columns:
            df_cruce[col] = pd.to_numeric(df_cruce[col], errors='coerce').fillna(0)
        else:
            df_cruce[col] = 0.0
            
    df_cruce['SUMA_CON_IVA'] = df_cruce['TotalProductosConIVa'] + df_cruce['TotalServiciosConIVa']
    df_cruce['SUMA_SIN_IVA'] = df_cruce['TotalProductosSinIVa'] + df_cruce['TotalServiciosSinIVa']
    
    # Check sorting columns
    sort_cols = ['MF_Lista2', 'INVOICE_DATE', 'DOC_REFERENCE']
    available_sort_cols = [c for c in sort_cols if c in df_cruce.columns]
    if available_sort_cols:
        df_cruce = df_cruce.sort_values(
            by=available_sort_cols,
            ascending=[True] * len(available_sort_cols)
        )
    
    df_cruce = df_cruce.head(100000)
    
    columnas_finales = [
        "INVOICE_DATE", "CORP", "VENDOR_ID", "VENDOR_NAME", "RUC_or_FED_ID",
        "MEMO", "INVOICE_TOTAL", "DOC_REFERENCE", "MF_Nume1", "MF_Alfa2",
        "MF_Lista2", "SUMA_CON_IVA", "SUMA_SIN_IVA", "ES_ANULADO", "MF_Bool5",
        "DOC_ID_CORP", "ID_Relacionado"
    ]
    
    columnas_disponibles = [col for col in columnas_finales if col in df_cruce.columns]
    df_final = df_cruce[columnas_disponibles]
    
    if not df_final.empty:
        df_final = df_final.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    return df_final

# ----------------- EXCEL GENERATION UTILITIES -----------------

def generate_excel_file(df, sheet_name):
    """
    Creates an Excel file in-memory matching the design of original scripts.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        
        fill_encabezado = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        fuente_encabezado = Font(color="FFFFFF", bold=True)
        alineacion_centrada = Alignment(horizontal="center", vertical="center")
        
        # Apply header styling
        for col_num in range(1, len(df.columns) + 1):
            celda = worksheet.cell(row=1, column=col_num)
            celda.fill = fill_encabezado
            celda.font = fuente_encabezado
            celda.alignment = alineacion_centrada
            
        # Adjust column widths with limit of 40
        for idx, col_name in enumerate(df.columns):
            ancho_calculado = max(
                df[col_name].astype(str).map(len).max(),
                len(str(col_name))
            ) + 3
            ancho_final = min(ancho_calculado, 40)
            letra_columna = get_column_letter(idx + 1)
            worksheet.column_dimensions[letra_columna].width = ancho_final
            
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        
    output.seek(0)
    return output

# ----------------- API ENDPOINTS -----------------

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/api/movimientos', methods=['GET'])
def get_movimientos():
    fecha_inicio = request.args.get('inicio')
    fecha_fin = request.args.get('fin')
    
    if not fecha_inicio or not fecha_fin:
        return jsonify({"error": "Faltan parámetros 'inicio' o 'fin'"}), 400
        
    token = obtener_token()
    if not token:
        return jsonify({"error": "No se pudo obtener el token de conexión con el ERP"}), 500
        
    df = get_movimientos_df(token, fecha_inicio, fecha_fin)
    return jsonify(df.to_dict(orient='records'))

@app.route('/api/liquidaciones', methods=['GET'])
def get_liquidaciones():
    fecha_inicio = request.args.get('inicio')
    fecha_fin = request.args.get('fin')
    
    if not fecha_inicio or not fecha_fin:
        return jsonify({"error": "Faltan parámetros 'inicio' o 'fin'"}), 400
        
    token = obtener_token()
    if not token:
        return jsonify({"error": "No se pudo obtener el token de conexión con el ERP"}), 500
        
    df = get_liquidaciones_df(token, fecha_inicio, fecha_fin)
    return jsonify(df.to_dict(orient='records'))

@app.route('/api/ats', methods=['GET'])
def get_ats():
    fecha_inicio = request.args.get('inicio')
    fecha_fin = request.args.get('fin')
    
    if not fecha_inicio or not fecha_fin:
        return jsonify({"error": "Faltan parámetros 'inicio' o 'fin'"}), 400
        
    token = obtener_token()
    if not token:
        return jsonify({"error": "No se pudo obtener el token de conexión con el ERP"}), 500
        
    df = get_ats_df(token, fecha_inicio, fecha_fin)
    return jsonify(df.to_dict(orient='records'))

@app.route('/api/excel/movimientos', methods=['GET'])
def download_excel_movimientos():
    fecha_inicio = request.args.get('inicio')
    fecha_fin = request.args.get('fin')
    
    if not fecha_inicio or not fecha_fin:
        return "Faltan parámetros 'inicio' o 'fin'", 400
        
    token = obtener_token()
    if not token:
        return "No se pudo conectar al ERP", 500
        
    df = get_movimientos_df(token, fecha_inicio, fecha_fin)
    if df.empty:
        return "No hay datos para exportar en este rango de fechas", 404
        
    excel_file = generate_excel_file(df, 'Movimientos')
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Reporte_Movimientos_MBA3_{timestamp}.xlsx"
    
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/excel/liquidaciones', methods=['GET'])
def download_excel_liquidaciones():
    fecha_inicio = request.args.get('inicio')
    fecha_fin = request.args.get('fin')
    
    if not fecha_inicio or not fecha_fin:
        return "Faltan parámetros 'inicio' o 'fin'", 400
        
    token = obtener_token()
    if not token:
        return "No se pudo conectar al ERP", 500
        
    df = get_liquidaciones_df(token, fecha_inicio, fecha_fin)
    if df.empty:
        return "No hay datos para exportar en este rango de fechas", 404
        
    excel_file = generate_excel_file(df, 'Consolidado')
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Reporte_Liquidaciones_Consolidado_{timestamp}.xlsx"
    
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/excel/ats', methods=['GET'])
def download_excel_ats():
    fecha_inicio = request.args.get('inicio')
    fecha_fin = request.args.get('fin')
    
    if not fecha_inicio or not fecha_fin:
        return "Faltan parámetros 'inicio' o 'fin'", 400
        
    token = obtener_token()
    if not token:
        return "No se pudo conectar al ERP", 500
        
    df = get_ats_df(token, fecha_inicio, fecha_fin)
    if df.empty:
        return "No hay datos para exportar en este rango de fechas", 404
        
    excel_file = generate_excel_file(df, 'Consolidado')
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Reporte_Facturacion_Fiscal_{timestamp}.xlsx"
    
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    logging.info("Iniciando Servidor del Dashboard de BI...")
    app.run(host='0.0.0.0', port=5000, debug=True)
