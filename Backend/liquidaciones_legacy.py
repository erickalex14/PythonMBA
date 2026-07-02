import requests
import logging
import pandas as pd
import json
import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configuracion del registro de eventos
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

BASE_URL = "http://181.198.104.181:8020"
URL_LOGIN = f"{BASE_URL}/ws2_mba3_serv_/login_servicio"
URL_CONSULTA = f"{BASE_URL}/ws2_mba3_serv_Consultas_Externas_/"

CODIGO_SERVICIO = "SERIALES"
PASSWORD_SERVICIO = "Admin2026@@"


def solicitar_fecha(mensaje_prompt):
    """
    Solicita al usuario una fecha por consola y valida que el formato sea YYYY-MM-DD.
    """
    while True:
        fecha_str = input(mensaje_prompt).strip()
        try:
            datetime.datetime.strptime(fecha_str, "%Y-%m-%d")
            return fecha_str
        except ValueError:
            print("Error: Formato de fecha incorrecto. Por favor, utilice el formato YYYY-MM-DD (ejemplo: 2026-01-15).")


def solicitar_rango_fechas():
    """
    Solicita y valida el rango de fechas, asegurando coherencia temporal.
    """
    print("\n--- Generador de Reporte de Liquidaciones Consolidado MBA3 ---")
    print("Por favor, ingrese el rango de fechas para la consulta.\n")

    while True:
        fecha_inicio = solicitar_fecha("Ingrese la fecha de INICIO (YYYY-MM-DD): ")
        fecha_fin = solicitar_fecha("Ingrese la fecha de FIN (YYYY-MM-DD): ")

        if fecha_inicio <= fecha_fin:
            print(f"\nRango seleccionado: Desde {fecha_inicio} hasta {fecha_fin}\n")
            return fecha_inicio, fecha_fin
        else:
            print("\nError: La fecha de inicio no puede ser posterior a la fecha de fin. Intente nuevamente.\n")


def obtener_token():
    logging.info("Iniciando proceso de autenticacion en MBA3.")
    headers = {"Content-Type": "application/json"}
    payload = {"codigo": CODIGO_SERVICIO, "pwd": PASSWORD_SERVICIO}

    try:
        response = requests.post(URL_LOGIN, json=payload, headers=headers, timeout=10)
        response.raise_for_status()
        datos = response.json()
        token = datos.get("jwt")
        if token:
            logging.info("Autenticacion exitosa. Token asignado.")
            return token
        else:
            logging.error("Respuesta del servidor sin token JWT: %s", datos)
            return None
    except requests.exceptions.RequestException as e:
        logging.error("Fallo de conexion en endpoint de login: %s", e)
        return None


def ejecutar_consulta_tabla(token, columnas, tabla, condicion_where=None, limite=None):
    headers = {"Authorization": token}
    payload = {"select": columnas, "from": tabla}

    if condicion_where:
        payload["where"] = condicion_where
    if limite:
        payload["limit"] = str(limite)

    try:
        response = requests.post(URL_CONSULTA, headers=headers, data=payload, timeout=90)
        response.raise_for_status()
        datos = response.json()

        if isinstance(datos, dict):
            if datos.get("codigo") == "009":
                return []
            elif "error" in datos or "codigo" in datos:
                logging.error("El servidor ERP rechazo la operacion. Detalle tecnico: %s", datos)
                return []
        return datos

    except requests.exceptions.RequestException as e:
        logging.error("Fallo de red al intentar consumir la tabla %s: %s", tabla, e)
        return []
    except json.JSONDecodeError:
        logging.error("Fallo en la decodificacion JSON para la tabla %s.", tabla)
        return []


def procesar_cruce_datos(token, fecha_inicio, fecha_fin):
    # 1. Extraccion de Tabla Principal con Filtro de Fechas Dinamico
    cols_principal = "CORP,LIQUIDACION_FECHA,OBSERVACIONES,ANTES_TOTAL_1,ANTES_TOTAL_2,ANTES_TOTAL_3,DESPUES_TOTAL_1,DESPUES_TOTAL_2,DESPUES_TOTAL_3,LIQUIDACION_ESTADO,LIQUIDACION_ID_CORP"

    # Inyeccion de las variables de fecha utilizando f-strings
    where_principal = f"CORP = 'NVC01' AND LIQUIDACION_FECHA >= '{fecha_inicio}' AND LIQUIDACION_FECHA <= '{fecha_fin}'"

    logging.info("Fase 1/3: Descargando datos de cabecera para el periodo especificado.")
    datos_principal = ejecutar_consulta_tabla(token, cols_principal, "PROV_Liquidaciones_Principal", where_principal,
                                              5000)

    if not datos_principal:
        logging.warning("Sin resultados en tabla principal para este rango de fechas. Proceso abortado.")
        return None

    df_principal = pd.DataFrame(datos_principal if isinstance(datos_principal, list) else [datos_principal])
    logging.info("Registros obtenidos en cabecera: %d", len(df_principal))

    # 2. Extraccion Masiva de Tabla Secundaria
    cols_productos = "FACTURA_ID_CORP,IdRecepcionRelacionada,VALOR_TOTAL_CIF,VALOR_SUBTOTAL_CIF,VALOR_ANTES_1,VALOR_ANTES_2,VALOR_ANTES_3,VALOR_DESPUES_1,VALOR_DESPUES_2,VALOR_DESPUES_3,PARTIDA_ID_CORP,PRODUCTO_ID_CORP,LIQUIDACION_ID,CANTIDAD,PRECIO,TOTAL,VALOR_TOTAL_CIF_MANUAL,VALOR_TOTAL_CIF_UNIDAD,LIQUIDACION_ID_CORP"

    logging.info("Fase 2/3: Descargando tabla de productos de forma MASIVA...")

    datos_productos = ejecutar_consulta_tabla(token, cols_productos, "PROV_Liquidaciones_Productos", limite=50000)

    if not datos_productos:
        logging.warning("Sin resultados en tabla de productos. El cruce carecera de detalles.")
        return None

    df_productos = pd.DataFrame(datos_productos if isinstance(datos_productos, list) else [datos_productos])
    logging.info("Registros obtenidos en productos: %d", len(df_productos))

    # 3. Operacion INNER JOIN en memoria RAM
    logging.info("Fase 3/3: Ejecutando cruce relacional (INNER JOIN) en memoria RAM.")
    df_cruce = pd.merge(df_principal, df_productos, on='LIQUIDACION_ID_CORP', how='inner')

    columnas_finales = [
        "CORP", "LIQUIDACION_FECHA", "OBSERVACIONES", "FACTURA_ID_CORP",
        "IdRecepcionRelacionada", "ANTES_TOTAL_1", "ANTES_TOTAL_2", "ANTES_TOTAL_3",
        "DESPUES_TOTAL_1", "DESPUES_TOTAL_2", "DESPUES_TOTAL_3", "VALOR_TOTAL_CIF",
        "VALOR_SUBTOTAL_CIF", "VALOR_ANTES_1", "VALOR_ANTES_2", "VALOR_ANTES_3",
        "VALOR_DESPUES_1", "VALOR_DESPUES_2", "VALOR_DESPUES_3", "PARTIDA_ID_CORP",
        "PRODUCTO_ID_CORP", "LIQUIDACION_ID", "CANTIDAD", "PRECIO", "TOTAL",
        "VALOR_TOTAL_CIF_MANUAL", "VALOR_TOTAL_CIF_UNIDAD", "LIQUIDACION_ESTADO",
        "LIQUIDACION_ID_CORP"
    ]

    columnas_disponibles = [col for col in columnas_finales if col in df_cruce.columns]
    df_final = df_cruce[columnas_disponibles]

    logging.info("Operacion de cruce finalizada. Lineas listas para exportacion: %d", len(df_final))
    return df_final


def exportar_a_excel_empresarial(df_datos, nombre_base="Reporte_Liquidaciones_Consolidado"):
    logging.info("Iniciando formateo de presentacion corporativa en hoja de calculo.")

    if df_datos is None or df_datos.empty:
        logging.warning("Estructura de datos nula. Operacion de guardado cancelada.")
        return

    try:
        df_datos = df_datos.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo_final = f"{nombre_base}_{timestamp}.xlsx"

        with pd.ExcelWriter(archivo_final, engine='openpyxl') as writer:
            df_datos.to_excel(writer, index=False, sheet_name='Consolidado')
            worksheet = writer.sheets['Consolidado']

            fill_encabezado = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            fuente_encabezado = Font(color="FFFFFF", bold=True)
            alineacion_centrada = Alignment(horizontal="center", vertical="center")

            for col_num in range(1, len(df_datos.columns) + 1):
                celda = worksheet.cell(row=1, column=col_num)
                celda.fill = fill_encabezado
                celda.font = fuente_encabezado
                celda.alignment = alineacion_centrada

            for idx, col_name in enumerate(df_datos.columns):
                ancho_calculado = max(
                    df_datos[col_name].astype(str).map(len).max(),
                    len(str(col_name))
                ) + 3
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(ancho_calculado, 40)

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

        logging.info("Secuencia concluida. Archivo de reporte consolidado: %s", archivo_final)

    except Exception as e:
        logging.error("Fallo critico en modulo de procesamiento de Excel: %s", e)


if __name__ == "__main__":
    # Solicitud inicial de parametros
    fecha_inicio, fecha_fin = solicitar_rango_fechas()

    token_jwt = obtener_token()

    if token_jwt:
        # Inyeccion de parametros hacia la logica principal
        df_resultado = procesar_cruce_datos(token_jwt, fecha_inicio, fecha_fin)
        exportar_a_excel_empresarial(df_resultado)
    else:
        logging.critical("Ejecucion detenida. Credenciales invalidas o error de red.")




