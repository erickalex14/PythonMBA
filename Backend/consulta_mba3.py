import requests
import logging
import pandas as pd
import json
import datetime
import time
import re
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configuracion del registro de eventos
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Constants de configuracion
BASE_URL = "http://181.198.104.181:8020"
URL_LOGIN = f"{BASE_URL}/ws2_mba3_serv_/login_servicio"
URL_CONSULTA = f"{BASE_URL}/ws2_mba3_serv_Consultas_Externas_/"

CODIGO_SERVICIO = "ERICKDEV"
PASSWORD_SERVICIO = "***REDACTED***"


def solicitar_fecha(mensaje_prompt):
    while True:
        fecha_str = input(mensaje_prompt).strip()
        try:
            fecha_obj = datetime.datetime.strptime(fecha_str, "%Y-%m-%d")
            return fecha_obj
        except ValueError:
            print("Error: Formato de fecha incorrecto. Utilice YYYY-MM-DD (ejemplo: 2026-06-01).")


def solicitar_rango_fechas():
    print("\n--- Generador Masivo de Reporte de Movimientos MBA3 ---")
    print("Por favor, ingrese el rango de fechas para la consulta.\n")

    while True:
        fecha_inicio = solicitar_fecha("Ingrese la fecha de INICIO (YYYY-MM-DD): ")
        fecha_fin = solicitar_fecha("Ingrese la fecha de FIN (YYYY-MM-DD): ")

        if fecha_inicio <= fecha_fin:
            print(
                f"\nRango seleccionado: Desde {fecha_inicio.strftime('%Y-%m-%d')} hasta {fecha_fin.strftime('%Y-%m-%d')}\n")
            return fecha_inicio, fecha_fin
        else:
            print("\nError: La fecha de inicio no puede ser posterior a la fecha de fin. Intente nuevamente.\n")


def obtener_token():
    logging.info("Solicitando nuevas credenciales de sesion al ERP...")
    headers = {"Content-Type": "application/json"}
    payload = {
        "codigo": CODIGO_SERVICIO,
        "pwd": PASSWORD_SERVICIO
    }

    try:
        response = requests.post(URL_LOGIN, json=payload, headers=headers, timeout=15)
        response.raise_for_status()

        datos = response.json()
        token = datos.get("jwt")

        if token:
            logging.info("Autenticacion exitosa. Nuevo token asignado.")
            return token
        else:
            logging.error("Respuesta de autenticacion sin token JWT. Detalle: %s", datos)
            return None

    except requests.exceptions.RequestException as e:
        logging.error("Fallo de red al intentar autenticar: %s", e)
        return None


def ejecutar_consulta_diaria(token, fecha_consulta):
    headers = {"Authorization": token}
    fecha_str = fecha_consulta.strftime('%Y-%m-%d')
    condicion_where = f"TRANS_DATE = '{fecha_str}'"
    columnas_select = "TRANS_DATE,PRODUCT_NAME,Codigo_producto_convertido,ORIGINAL_QTY,ORIGIN_MEMO,ORIGIN_REF,BASE_COMISION,Info_Seriales,Codigo_Sucursal,BaseImponibleReal_1,COD_SALESMAN,Codigo_Marca"

    payload = {
        "select": columnas_select,
        "from": "INVT_Producto_Movimientos",
        "where": condicion_where,
        "limit": "50000"
    }

    try:
        response = requests.post(URL_CONSULTA, headers=headers, data=payload, timeout=150)

        if response.status_code == 401:
            return "UNAUTHORIZED"

        response.raise_for_status()
        datos = response.json()

        if isinstance(datos, dict):
            if datos.get("codigo") == "009":
                return []
            elif "error" in datos or "codigo" in datos:
                logging.error("El servidor rechazo la consulta. Detalle: %s", datos)
                return []

        if isinstance(datos, dict):
            return [datos]
        elif isinstance(datos, list):
            return datos
        else:
            return []

    except requests.exceptions.Timeout:
        return "TIMEOUT"
    except requests.exceptions.RequestException as e:
        logging.error("Error de conexion general: %s", e)
        return "ERROR"
    except json.JSONDecodeError:
        logging.error("Fallo en la decodificacion JSON.")
        return "ERROR"


def procesar_extraccion_masiva(token_inicial, fecha_inicio, fecha_fin):
    registros_totales = []
    fecha_actual = fecha_inicio
    dias_totales = (fecha_fin - fecha_inicio).days + 1
    dia_contador = 1
    token_actual = token_inicial

    logging.info("Iniciando extraccion masiva. Total de dias a procesar: %d", dias_totales)

    while fecha_actual <= fecha_fin:
        fecha_str = fecha_actual.strftime('%Y-%m-%d')
        logging.info("Procesando lote %d/%d (Fecha: %s)", dia_contador, dias_totales, fecha_str)

        max_intentos = 3
        intento_actual = 1
        exito = False

        while intento_actual <= max_intentos and not exito:
            resultado = ejecutar_consulta_diaria(token_actual, fecha_actual)

            if resultado == "UNAUTHORIZED":
                logging.warning("La sesion expiro. Renovando token automaticamente...")
                nuevo_token = obtener_token()
                if nuevo_token:
                    token_actual = nuevo_token
                    time.sleep(1)
                else:
                    logging.critical("Fallo critico al renovar credenciales. Abortando extraccion.")
                    return pd.DataFrame(registros_totales)

            elif resultado == "TIMEOUT":
                logging.warning("Timeout detectado para el dia %s. Reintentando (Intento %d/%d)...", fecha_str,
                                intento_actual, max_intentos)
                intento_actual += 1
                time.sleep(5)

            elif resultado == "ERROR":
                logging.error("Error estructural en el lote del %s. Se omitira este dia.", fecha_str)
                break

            else:
                if resultado:
                    registros_totales.extend(resultado)
                exito = True

        if not exito and intento_actual > max_intentos:
            logging.error("Se agotaron los reintentos para el dia %s. Datos omitidos.", fecha_str)

        time.sleep(1.0)
        fecha_actual += datetime.timedelta(days=1)
        dia_contador += 1

    if not registros_totales:
        logging.warning("El proceso finalizo sin recuperar registros.")
        return None

    df_resultado = pd.DataFrame(registros_totales)
    logging.info("Extraccion completada exitosamente. Total de registros recuperados: %d", len(df_resultado))
    return df_resultado


def exportar_a_excel_empresarial(df_datos, fecha_inicio, fecha_fin):
    """
    Exporta el DataFrame a un archivo Excel con formatos empresariales estandarizados.
    El nombre del archivo se genera dinámicamente con el rango de fechas.
    """
    logging.info("Iniciando modulo de renderizado de reporte Excel corporativo con saneamiento de datos.")

    if df_datos is None or df_datos.empty:
        logging.warning("Conjunto de datos vacio. Se omite la generacion del archivo.")
        return

    try:
        # Expresion regular para detectar caracteres de control ilegales en archivos Excel (XML)
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

        def sanitizar_excel(valor):
            if pd.isna(valor):
                return valor
            val_str = str(valor).strip()
            val_str = ILLEGAL_CHARACTERS_RE.sub("", val_str)
            if len(val_str) > 32700:
                return val_str[:32700] + "... [TRUNCADO POR LIMITE DE EXCEL]"
            return val_str

        # Aplicamos la limpieza extrema
        for col in df_datos.select_dtypes(include=['object']).columns:
            df_datos[col] = df_datos[col].apply(sanitizar_excel)

        # Generacion del nombre dinamico: SERIALES (YYYY-MM-DD al YYYY-MM-DD).xlsx
        str_inicio = fecha_inicio.strftime('%Y-%m-%d')
        str_fin = fecha_fin.strftime('%Y-%m-%d')
        archivo_final = f"SERIALES ({str_inicio} al {str_fin}).xlsx"

        with pd.ExcelWriter(archivo_final, engine='openpyxl') as writer:
            df_datos.to_excel(writer, index=False, sheet_name='Movimientos')
            worksheet = writer.sheets['Movimientos']

            fill_encabezado = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            fuente_encabezado = Font(color="FFFFFF", bold=True)
            alineacion_centrada = Alignment(horizontal="center", vertical="center")

            # Aplicar estilos de encabezado
            for col_num in range(1, len(df_datos.columns) + 1):
                celda = worksheet.cell(row=1, column=col_num)
                celda.fill = fill_encabezado
                celda.font = fuente_encabezado
                celda.alignment = alineacion_centrada

            # Ajuste de ancho de columnas
            for idx, col_name in enumerate(df_datos.columns):
                ancho_calculado = max(
                    df_datos[col_name].astype(str).map(len).max(),
                    len(str(col_name))
                ) + 3
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(ancho_calculado, 40)

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

        logging.info("Renderizado exitoso. Archivo generado: %s", archivo_final)

    except Exception as e:
        logging.error("Fallo critico I/O durante la escritura del archivo Excel: %s", e)


if __name__ == "__main__":
    fecha_inicio, fecha_fin = solicitar_rango_fechas()

    token_jwt = obtener_token()

    if token_jwt:
        df_resultado = procesar_extraccion_masiva(token_jwt, fecha_inicio, fecha_fin)
        exportar_a_excel_empresarial(df_resultado, fecha_inicio, fecha_fin)
    else:
        logging.critical("Secuencia interrumpida. No fue posible establecer conexion.")