"""Valida la agregacion de KPIs contra el Excel de Seguimiento.

Replica lo que hace el SQL de `KpiService` (cruzar el catalogo producto->categoria
y agrupar por sucursal, en unidades o en dinero segun el KPI) pero sobre la hoja
BASE del Excel, y compara el resultado celda por celda contra los valores reales
de la hoja RESUMEN KPI.

Esto prueba la LOGICA, no la consulta: que el SQL devuelva las mismas filas que
la hoja BASE depende del sync, y eso se verifica corriendo el endpoint.

Correr con:
    py -3 Backend/scripts/test_kpi_agregacion.py "C:/ruta/SEGUIMIENTO KPI.xlsx"
"""
import collections
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from app.services.kpi_service import CATS_VENTAS, KPIS  # noqa: E402

# Columna de RESUMEN KPI con el valor REAL de cada KPI (0-indexado).
COL_REAL = {5: "tecnologia", 7: "celulares_tablets", 9: "motorola",
            11: "sillas_gamer", 13: "hogar_gym", 21: "servicio_tecnico"}

RE_SUCURSAL = re.compile(r"^\s*(\d{3})")
TOL = 0.02      # el Excel redondea a centavos


def main():
    ruta = sys.argv[1] if len(sys.argv) > 1 else None
    if not ruta:
        print(__doc__)
        raise SystemExit(2)

    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)

    cat = {}
    for r in wb["POND"].iter_rows(min_row=2, min_col=10, max_col=12, values_only=True):
        if r[0] and r[2]:
            cat[str(r[0]).strip().upper()] = str(r[2]).strip().upper()

    cat_a_kpi = {v.upper().strip(): k for k, v in CATS_VENTAS.items()}
    unidades, monto = collections.Counter(), collections.Counter()
    for r in wb["BASE"].iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        m = RE_SUCURSAL.match(str(r[6] or ""))
        kpi = cat_a_kpi.get(cat.get(str(r[8] or "").strip().upper(), ""))
        if not (m and kpi):
            continue
        unidades[(m.group(1), kpi)] += r[13] or 0
        monto[(m.group(1), kpi)] += r[15] or 0

    ok, fallos = 0, []
    for row in wb["RESUMEN KPI"].iter_rows(min_row=3, values_only=True):
        m = RE_SUCURSAL.match(str(row[1] or ""))
        if not m:
            continue
        for col, kpi in COL_REAL.items():
            esperado = row[col] if col < len(row) else None
            if not isinstance(esperado, (int, float)):
                continue
            fuente = monto if KPIS[kpi]["medida"] == "monto" else unidades
            got = float(fuente[(m.group(1), kpi)])
            if abs(got - float(esperado)) < TOL:
                ok += 1
            else:
                fallos.append((m.group(1), kpi, float(esperado), got))

    total = ok + len(fallos)
    print(f"coinciden {ok} de {total} celdas ({ok / total * 100:.1f}%)")
    for s, k, e, g in fallos[:10]:
        print(f"  {s} {k:20} excel={e:10.2f} calculado={g:10.2f}")
    assert not fallos, f"{len(fallos)} celdas no cuadran contra el Excel"
    print("\nLa agregacion reproduce el Excel exactamente.")


if __name__ == "__main__":
    main()
