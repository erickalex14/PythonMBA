"""Verifica que el libro del Seguimiento KPI salga con todas sus hojas.

Arma el Excel con datos sinteticos (no toca la base) y revisa que existan las
hojas esperadas, que el detalle por categoria filtre bien y que BASE conserve
las lineas sin categoria.

Correr con:  py -3 Backend/scripts/test_kpi_excel.py
"""
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services.excel_service import ExcelService          # noqa: E402
from app.services.kpi_service import (COLUMNAS_PRESUPUESTO,  # noqa: E402
                                      COLUMNAS_RESUMEN, HOJA_POR_KPI, KPIS)

SEGUIMIENTO = {
    "periodo": "2026-08", "inicio": "2026-08-01", "corte": "2026-08-16",
    "dias_corte": 16, "dias_mes": 31, "peso_total": 0.30,
    "sucursales": [{
        "sucursal": "001", "nombre": "RIO COCA", "supervisor": "GUSTAVO ONA",
        "marca": "NOVICOMPU", "ciudad": "QUITO",
        "total_kpi": 0.0935784, "sin_metas": False,
        "detalle": [{"kpi": k, "label": c["label"], "origen": c["origen"],
                     "medida": c["medida"], "peso": c["peso"], "meta": 10,
                     "real": 5, "cumplimiento": 0.5,
                     "aporte": c["peso"] * 0.5, "proyeccion": 9.7}
                    for k, c in KPIS.items()],
    }],
}

PRESUPUESTO = [{
    "sucursal": "001", "nombre": "RIO COCA", "supervisor": "GUSTAVO ONA",
    "marca": "NOVICOMPU", "ciudad": "QUITO", "meta": 50000.0,
    "venta": 16331.6, "facturas": 159, "unidades": 254.0,
    "ticket_promedio": 102.71, "unidades_por_factura": 1.6,
    "venta_promedio_dia": 1020.72, "proyeccion": 31642.47, "cumplimiento": 0.3266,
}]


def _linea(codigo, cat, total):
    return {"factura_final": "5001", "fecha": "2026-08-01", "codigo": codigo,
            "producto": f"PROD {codigo}", "unidad": "UN", "grupo": "ET",
            "subgrupo": "LAPTO", "cantidad": 1, "precio_venta": total,
            "total_linea": total, "bodega_codigo": "5", "sucursal": "001",
            "sucursal_nombre": "RIO COCA", "supervisor": "GUSTAVO ONA", "cat": cat,
            "codigo_vendedor": "apjc", "nombre_cliente": "CLIENTE X",
            "canal": "TIENDA", "sucursal_larga": "001 RIO COCA"}


LINEAS = [
    _linea("1EENV9220", "TECNOLOGIA", 533.57),
    _linea("1EENV9212", "TECNOLOGIA", 574.55),
    _linea("1CENV89", "CELULARES Y TABLETS ENV", 146.91),
    _linea("1HENV4565", "HOGAR Y GYM", 20.0),
    _linea("B2P-SERTEC1", "ST", 13.04),
    _linea("1CSAM5797", None, 152.0),      # marca de tercero: no puntua
]


def main():
    wb_bytes = ExcelService().generar_reporte_kpi(SEGUIMIENTO, PRESUPUESTO, LINEAS)
    wb = openpyxl.load_workbook(wb_bytes)

    esperadas = ["RESUMEN KPI", "PRESUPUESTO", *HOJA_POR_KPI.values(), "BASE"]
    faltan = [h for h in esperadas if h not in wb.sheetnames]
    assert not faltan, f"faltan hojas: {faltan}"
    print(f"OK {len(wb.sheetnames)} hojas: {', '.join(wb.sheetnames)}")

    def filas_datos(hoja):
        """Filas de datos: el encabezado va en la fila 1, como en el original."""
        ws = wb[hoja]
        return [r for r in ws.iter_rows(min_row=2, values_only=True)
                if r and str(r[0] or "").strip() == "5001"]

    assert len(filas_datos("TCNLG")) == 2, "TCNLG deberia traer las 2 de TECNOLOGIA"
    assert len(filas_datos("ST")) == 1, "ST deberia traer 1 linea"
    assert len(filas_datos("RAZR-EDGE60")) == 0, "no hay ventas de Motorola"
    print("OK el detalle por categoria filtra correctamente")

    assert len(filas_datos("BASE")) == len(LINEAS), \
        "BASE debe incluir tambien las lineas sin categoria"
    print(f"OK BASE conserva las {len(LINEAS)} lineas, incluida la no categorizada")

    # El encabezado tiene que ser IDENTICO al del archivo que arma Contabilidad,
    # espacios de mas incluidos: el reporte se compara mes a mes contra el suyo.
    ws = wb["RESUMEN KPI"]
    r1 = [str(c or "") for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    r2 = [c for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]

    assert r1[1] == "CORTE AL 2026-08-16", r1[1]
    assert r1[2] == "SUPERVISOR" and r2[1] == "SUCURSAL", (r1[2], r2[1])
    for kpi, tit_real, tit_meta in COLUMNAS_RESUMEN:
        assert tit_real in r1, f"falta la columna real {tit_real!r}"
        assert tit_meta in r1, f"falta la columna meta {tit_meta!r}"
        # El peso va justo debajo del titulo de la meta.
        assert r2[r1.index(tit_meta)] == KPIS[kpi]["peso"], kpi
    assert r1[-1] == "% CUMPLIMIENTO KPI" and r2[-1] == "total", (r1[-1], r2[-1])
    print("OK RESUMEN KPI replica los encabezados del archivo manual")

    # La primera sucursal arranca en la fila 3, con el codigo pegado al nombre.
    fila3 = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    assert fila3[0] == 1 and fila3[1] == "001 RIO COCA", fila3[:2]
    print("OK los datos empiezan en la fila 3, como en el original")

    cab_pre = [str(c or "") for c in
               next(wb["PRESUPUESTO"].iter_rows(min_row=1, max_row=1, values_only=True))]
    assert cab_pre[:6] == COLUMNAS_PRESUPUESTO[:6], cab_pre[:6]
    print("OK PRESUPUESTO replica sus encabezados")

    cab_base = [str(c or "") for c in
                next(wb["BASE"].iter_rows(min_row=1, max_row=1, values_only=True))]
    assert cab_base[:7] == ["No. Factura", "Código", "Nombre", "CANAL", "Fecha",
                            "Bodega", "SUCURSAL"], cab_base[:7]
    print("OK las hojas de detalle replican sus encabezados")

    print("\nEl libro sale con el formato del archivo manual.")


if __name__ == "__main__":
    main()
