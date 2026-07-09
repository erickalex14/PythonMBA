import requests
import logging
import pandas as pd
import json
import datetime
import re
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configuracion del registro de eventos profesional
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Constantes de configuracion para produccion
BASE_URL = "http://181.198.104.181:8020"
URL_LOGIN = f"{BASE_URL}/ws2_mba3_serv_/login_servicio"
URL_CONSULTA = f"{BASE_URL}/ws2_mba3_serv_Consultas_Externas_/"

CODIGO_SERVICIO = "ERICKDEV"
PASSWORD_SERVICIO = "Er1ck2026$$"


def solicitar_fecha_reporte():
    print("\n--- Generador de Reporte Relacional Espejo Completo (MBA3) ---")
    while True:
        fecha_str = input("Ingrese la fecha a consultar (YYYY-MM-DD): ").strip()
        try:
            fecha_obj = datetime.datetime.strptime(fecha_str, "%Y-%m-%d")
            return fecha_obj
        except ValueError:
            print("Error: Formato incorrecto. Utilice YYYY-MM-DD (ejemplo: 2026-06-01).")


def obtener_token():
    logging.info("Solicitando credenciales de sesion al ERP...")
    headers = {"Content-Type": "application/json"}
    payload = {"codigo": CODIGO_SERVICIO, "pwd": PASSWORD_SERVICIO}

    try:
        requests_response = requests.post(URL_LOGIN, json=payload, headers=headers, timeout=15)
        requests_response.raise_for_status()
        datos = requests_response.json()
        token = datos.get("jwt")
        if token:
            logging.info("Autenticacion exitosa. Token asignado.")
            return token
        return None
    except Exception as e:
        logging.error("Fallo de red al autenticar: %s", e)
        return None


def ejecutar_consulta_tabla(token, columnas, tabla, condicion_where):
    headers = {"Authorization": token}
    payload = {
        "select": columnas,
        "from": tabla,
        "where": condicion_where,
        "limit": "300000"
    }
    try:
        response = requests.post(URL_CONSULTA, headers=headers, data=payload, timeout=240)
        response.raise_for_status()
        return response.json()
    except Exception as e:
        logging.error("Error al consultar la tabla %s: %s", tabla, e)
        return []


def generar_reporte_espejo_total(token, fecha_consulta):
    fecha_str = fecha_consulta.strftime('%Y-%m-%d')

    # 1. EXTRACCIÓN DEL KARDEX DE INVENTARIO
    logging.info("Paso 1/2: Descargando transacciones de inventario del Kardex...")
    cols_movs = (
        "DOC_ID_CORP,TRANS_DATE,PRODUCT_ID_CORP,PRODUCT_NAME,QUANTITY,ORIGINAL_QTY,"
        "UNIT_COST,DISCOUNT_AMOUNT,NET_LINE_TOTAL,UM,Anulada,IN_OUT,"
        "\"Codigo grupo\",\"Codigo subgrupo\",Codigo_grupo,Codigo_subgrupo"
    )
    where_movs = f"TRANS_DATE = '{fecha_str}'"
    datos_movs = ejecutar_consulta_tabla(token, cols_movs, "INVT_Producto_Movimientos", where_movs)

    if not datos_movs or (isinstance(datos_movs, dict) and "codigo" in datos_movs):
        logging.warning("No se encontraron movimientos de inventario para este dia.")
        return None

    df_movs = pd.DataFrame(datos_movs if isinstance(datos_movs, list) else [datos_movs])

    # 2. EXTRACCIÓN DE CABECERAS DE FACTURAS DE CLIENTES (CORREGIDO AL ORIGEN REAL)
    logging.info("Paso 2/2: Descargando cabeceras comerciales desde CLNT_Factura_Principal...")
    cols_facturas = "DOC_ID_CORP,DOC_REFERENCE,INVOICE_DATE"
    where_facturas = f"INVOICE_DATE = '{fecha_str}'"
    datos_facturas = ejecutar_consulta_tabla(token, cols_facturas, "CLNT_Factura_Principal", where_facturas)

    df_facturas = pd.DataFrame(datos_facturas if isinstance(datos_facturas, list) else [datos_facturas])

    if df_facturas.empty:
        logging.warning("La tabla CLNT_Factura_Principal no retorno datos. Heredando identificador base del Kardex.")
        df_movs['FACTURA_FINAL'] = df_movs['DOC_ID_CORP']
        df_consolidado = df_movs.copy()
    else:
        # Normalizamos mapeos contra las variantes de 4D
        mapeo_movs = {c.replace(" ", "").replace("_", "").upper(): c for c in df_movs.columns}
        col_movs_doc = mapeo_movs.get("DOCIDCORP")
        col_codigo_prod = mapeo_movs.get("PRODUCTIDCORP")
        col_nombre_prod = mapeo_movs.get("PRODUCTNAME")
        col_cant_real = mapeo_movs.get("ORIGINALQTY") if mapeo_movs.get("ORIGINALQTY") else mapeo_movs.get("QUANTITY")
        col_grupo = mapeo_movs.get("CODIGOGRUPO")
        col_subgrupo = mapeo_movs.get("CODIGOSUBGRUPO")
        col_precio = mapeo_movs.get("UNITCOST")
        col_descuento = mapeo_movs.get("DISCOUNTAMOUNT")
        col_total_linea = mapeo_movs.get("NETLINETOTAL")

        df_movs = df_movs.rename(columns={
            col_movs_doc: 'DOC_ID_CORP_KARDEX',
            col_codigo_prod: 'CODIGO_INT',
            col_nombre_prod: 'PRODUCTO_INT',
            col_cant_real: 'CANTIDAD_INT',
            col_grupo: 'GRUPO_INT',
            col_subgrupo: 'SUBGRUPO_INT',
            col_precio: 'PRECIO_INT',
            col_descuento: 'DESCUENTO_INT',
            col_total_linea: 'TOTAL_INT'
        })

        mapeo_fact = {c.replace(" ", "").replace("_", "").upper(): c for c in df_facturas.columns}
        col_fact_id = mapeo_fact.get("DOCIDCORP")
        col_fact_ref = mapeo_fact.get("DOCREFERENCE")

        df_facturas = df_facturas.rename(columns={
            col_fact_id: 'DOC_ID_CORP_FACT',
            col_fact_ref: 'NUMERO_FACTURA_REAL'
        })

        # Higiene de llaves relacionales extraídas por Regex numérico
        def limpiar_llave_numerica(val):
            if pd.isna(val): return ""
            numeros = re.findall(r'\d+', str(val))
            return "".join(numeros) if numeros else str(val).strip()

        df_movs['KEY_CRUCE_KARDEX'] = df_movs['DOC_ID_CORP_KARDEX'].apply(limpiar_llave_numerica)
        df_facturas['KEY_CRUCE_FACT'] = df_facturas['DOC_ID_CORP_FACT'].apply(limpiar_llave_numerica)

        logging.info("Cruzando Kardex con Facturas de Clientes en memoria RAM...")
        df_consolidado = pd.merge(df_movs, df_facturas, left_on='KEY_CRUCE_KARDEX', right_on='KEY_CRUCE_FACT',
                                  how='left')

        # Asignación del número de factura secuencial real
        df_consolidado['FACTURA_FINAL'] = df_consolidado['NUMERO_FACTURA_REAL'].fillna(
            df_consolidado['DOC_ID_CORP_KARDEX'])

        # Control preventivo contra variables vacías latentes
        df_consolidado['FACTURA_FINAL'] = df_consolidado.apply(
            lambda r: r['DOC_ID_CORP_KARDEX'] if str(r['FACTURA_FINAL']).strip().upper() in ['NAN', 'NONE', ''] else r[
                'FACTURA_FINAL'],
            axis=1
        )

    # 4. SANEAMIENTO Y FILTROS DE NEGOCIO
    df_consolidado['FACTURA_FINAL'] = df_consolidado['FACTURA_FINAL'].astype(str).str.replace(r'\.0$', '',
                                                                                              regex=True).str.strip().str.upper()
    df_consolidado['CODIGO_INT'] = df_consolidado['CODIGO_INT'].astype(str).str.replace(r'\.0$', '',
                                                                                        regex=True).str.strip().str.upper()
    df_consolidado['PRODUCTO_INT'] = df_consolidado['PRODUCTO_INT'].astype(str).str.strip().str.upper()
    df_consolidado['UM'] = df_consolidado['UM'].astype(
        str).str.strip().str.upper() if 'UM' in df_consolidado.columns else 'UNID'

    df_consolidado['CANTIDAD_INT'] = pd.to_numeric(df_consolidado['CANTIDAD_INT'], errors='coerce').fillna(0)
    df_consolidado['PRECIO_INT'] = pd.to_numeric(df_consolidado['PRECIO_INT'], errors='coerce').fillna(0.0)
    df_consolidado['DESCUENTO_INT'] = pd.to_numeric(df_consolidado['DESCUENTO_INT'], errors='coerce').fillna(0.0)
    df_consolidado['TOTAL_INT'] = pd.to_numeric(df_consolidado['TOTAL_INT'], errors='coerce').fillna(0.0)

    def evaluar_anulada(val):
        if pd.isna(val): return False
        if isinstance(val, bool): return val
        return str(val).strip().lower() in ['true', '1', 't', 's', 'si', 'y']

    df_consolidated_anulada = df_consolidado.copy()
    df_consolidated_anulada['IS_ANULADA'] = df_consolidated_anulada['Anulada'].apply(
        evaluar_anulada) if 'Anulada' in df_consolidated_anulada.columns else False

    df_filtrado = df_consolidated_anulada[
        (df_consolidated_anulada['IS_ANULADA'] == False) &
        (df_consolidated_anulada['CANTIDAD_INT'] > 0)
        ].copy()

    if df_filtrado.empty:
        logging.warning("Ninguna linea transaccional paso los filtros de auditoria de ventas.")
        return None

    df_filtrado['CANTIDAD_INT'] = df_filtrado['CANTIDAD_INT'].round(0).astype(int)
    df_filtrado['SUBTOTAL_INT'] = df_filtrado['CANTIDAD_INT'] * df_filtrado['PRECIO_INT']
    df_filtrado['TOTAL_INT'] = df_filtrado.apply(
        lambda r: r['TOTAL_INT'] if r['TOTAL_INT'] > 0 else (r['SUBTOTAL_INT'] - r['DESCUENTO_INT']), axis=1
    )

    df_espejo = pd.DataFrame()
    df_espejo['# de factura'] = df_filtrado['FACTURA_FINAL']
    df_espejo['FECHA'] = df_filtrado['TRANS_DATE']
    df_espejo['CODIGO'] = df_filtrado['CODIGO_INT']
    df_espejo['PRODUCTO'] = df_filtrado['PRODUCTO_INT']
    df_espejo['GRUPO'] = df_filtrado['GRUPO_INT'].fillna('GENERAL')
    df_espejo['SUBGRUPO'] = df_filtrado['SUBGRUPO_INT'].fillna('GENERAL')
    df_espejo['UNIDAD'] = df_filtrado['UM']
    df_espejo['CANTIDAD'] = df_filtrado['CANTIDAD_INT']
    df_espejo['PRECIO VENTA'] = df_filtrado['PRECIO_INT'].round(4)
    df_espejo['SUBTOTAL (C*PV)'] = df_filtrado['SUBTOTAL_INT'].round(4)
    df_espejo['DESCUENTO APLICADO'] = df_filtrado['DESCUENTO_INT'].round(4)
    df_espejo['TOTAL LINEA'] = df_filtrado['TOTAL_INT'].round(4)

    df_espejo = df_espejo.sort_values(by=['# de factura', 'CODIGO'], ascending=[True, True])
    return df_espejo


def exportar_a_excel_empresarial(df_datos, fecha_consulta):
    if df_datos is None or df_datos.empty:
        logging.warning("El conjunto de datos esta vacio. Se cancela la exportacion.")
        return

    logging.info("Generando renderizado final en archivo Excel estructurado...")
    try:
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

        def sanitizar_excel(valor):
            if pd.isna(valor): return valor
            return ILLEGAL_CHARACTERS_RE.sub("", str(valor).strip())

        for col in df_datos.select_dtypes(include=['object']).columns:
            df_datos[col] = df_datos[col].apply(sanitizar_excel)

        fecha_str = fecha_consulta.strftime('%Y-%m-%d')
        archivo_final = f"DETALLE_VENTAS_ESPEJO_({fecha_str}).xlsx"

        with pd.ExcelWriter(archivo_final, engine='openpyxl') as writer:
            df_datos.to_excel(writer, index=False, sheet_name='Detalle Productos - Periodo')
            worksheet = writer.sheets['Detalle Productos - Periodo']

            fill_encabezado = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            fuente_encabezado = Font(color="FFFFFF", bold=True)
            alineacion_centrada = Alignment(horizontal="center", vertical="center")

            for col_num in range(1, len(df_datos.columns) + 1):
                celda = worksheet.cell(row=1, column=col_num)
                celda.fill = fill_encabezado
                celda.font = fuente_encabezado
                celda.alignment = alineacion_centrada

            for idx, col_name in enumerate(df_datos.columns):
                ancho_calculado = max(df_datos[col_name].astype(str).map(len).max(), len(str(col_name))) + 4
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(ancho_calculado, 50)

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

        logging.info("¡REPORTE COMPLETADO AL 100%%! Archivo generado: %s", archivo_final)

    except Exception as e:
        logging.error("Fallo critico I/O al escribir el archivo Excel: %s", e)


if __name__ == "__main__":
    fecha_reporte = solicitar_fecha_reporte()
    token_jwt = obtener_token()

    if token_jwt:
        df_final = generar_reporte_espejo_total(token_jwt, fecha_reporte)
        exportar_a_excel_empresarial(df_final, fecha_reporte)
    else:
        logging.critical("Ejecucion abortada por falta de credenciales autorizadas.")
)