import os
import requests
import logging
import pandas as pd
import json
import datetime
import re
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Cargar variables de entorno
load_dotenv()

# Configuracion del registro de eventos
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

BASE_URL = os.getenv("MBA3_BASE_URL", "")
URL_LOGIN = f"{BASE_URL}/ws2_mba3_serv_/login_servicio"
URL_CONSULTA = f"{BASE_URL}/ws2_mba3_serv_Consultas_Externas_/"

CODIGO_SERVICIO = os.getenv("MBA3_CODIGO_SERVICIO", "")
PASSWORD_SERVICIO = os.getenv("MBA3_PASSWORD_SERVICIO", "")


def solicitar_fecha(mensaje_prompt):
    while True:
        fecha_str = input(mensaje_prompt).strip()
        try:
            datetime.datetime.strptime(fecha_str, "%Y-%m-%d")
            return fecha_str
        except ValueError:
            print("Error: Formato de fecha incorrecto. Utilice YYYY-MM-DD.")


def solicitar_rango_fechas():
    print("\n--- Generador de Reporte ATS (Anexo Transaccional Simplificado) ---")
    print("Por favor, ingrese el rango de fechas del periodo a declarar.\n")

    while True:
        fecha_inicio = solicitar_fecha("Ingrese la fecha de INICIO (YYYY-MM-DD): ")
        fecha_fin = solicitar_fecha("Ingrese la fecha de FIN (YYYY-MM-DD): ")

        if fecha_inicio <= fecha_fin:
            print(f"\nPeriodo seleccionado: Desde {fecha_inicio} hasta {fecha_fin}\n")
            return fecha_inicio, fecha_fin
        else:
            print("\nError: La fecha de inicio no puede ser posterior a la fecha de fin.\n")


def obtener_token():
    logging.info("Iniciando proceso de autenticacion en servidor ERP.")
    headers = {"Content-Type": "application/json"}
    payload = {"codigo": CODIGO_SERVICIO, "pwd": PASSWORD_SERVICIO}

    try:
        response = requests.post(URL_LOGIN, json=payload, headers=headers, timeout=15)
        response.raise_for_status()
        datos = response.json()
        token = datos.get("jwt")
        if token:
            logging.info("Autenticacion exitosa. Token asignado.")
            return token
        else:
            logging.error("Respuesta del servidor sin token JWT: %s", datos)
            return None
    except requests.exceptions.RequestException as e:
        logging.error("Fallo de conexion en endpoint de login: %s", e)
        return None


def ejecutar_consulta_tabla(token, columnas, tabla, de_atras_para_adelante=False, condicion_where=None, limite=None):
    headers = {"Authorization": token}
    payload = {"select": columnas, "from": tabla}

    if condicion_where:
        payload["where"] = condicion_where
    if limite:
        payload["limit"] = str(limite)

    # Si se activa, forzamos al motor a ordenar de forma descendente para capturar la data mas nueva
    if de_atras_para_adelante:
        payload["order_by"] = "ID_Relacionado DESC"

    try:
        response = requests.post(URL_CONSULTA, headers=headers, data=payload, timeout=300)

        if response.status_code == 401:
            return "UNAUTHORIZED"

        response.raise_for_status()
        datos = response.json()

        if isinstance(datos, dict):
            if datos.get("codigo") == "009":
                return []
            elif "error" in datos or "codigo" in datos:
                return "ERROR"
        return datos

    except Exception:
        return "ERROR"


def procesar_cruce_ats(token_inicial, fecha_inicio, fecha_fin):
    token_actual = token_inicial

    # =========================================================
    # 1. EXTRACCION DE CABECERAS (SQL: PROV_Factura_Principal)
    # =========================================================
    cols_factura = "INVOICE_DATE,CORP,VENDOR_ID,MEMO,INVOICE_TOTAL,DOC_REFERENCE,TotalProductosConIVa,TotalServiciosConIVa,TotalProductosSinIVa,TotalServiciosSinIVa,VOID,DOC_ID_CORP,CONFIRMED"
    where_factura = f"INVOICE_DATE >= '{fecha_inicio}' AND INVOICE_DATE <= '{fecha_fin}' AND CONFIRMED = True AND VOID = False"

    logging.info("Fase 1/4: Descargando cabeceras de facturas para el periodo del ATS.")
    datos_factura = ejecutar_consulta_tabla(token_actual, cols_factura, "PROV_Factura_Principal", False, where_factura,
                                            100000)

    if datos_factura == "UNAUTHORIZED" or datos_factura == "ERROR" or not datos_factura:
        logging.warning("Sin facturas validas registradas en este periodo. Abortando.")
        return None

    df_facturas = pd.DataFrame(datos_factura if isinstance(datos_factura, list) else [datos_factura])
    logging.info("Facturas base encontradas: %d", len(df_facturas))

    df_facturas['VENDOR_ID'] = df_facturas['VENDOR_ID'].astype(str).str.replace(r'\.0$', '',
                                                                                regex=True).str.strip().str.upper()
    df_facturas['DOC_ID_CORP'] = df_facturas['DOC_ID_CORP'].astype(str).str.replace(r'\.0$', '',
                                                                                    regex=True).str.strip().str.upper()

    # =========================================================
    # 2. EXTRACCION DE PROVEEDORES (SQL: PROV_Ficha_Principal)
    # =========================================================
    cols_proveedor = "VENDOR_ID,VENDOR_NAME,RUC_or_FED_ID"
    logging.info("Fase 2/4: Descargando y depurando catalogo de proveedores...")
    datos_proveedor = ejecutar_consulta_tabla(token_actual, cols_proveedor, "PROV_Ficha_Principal", False,
                                              limite=150000)
    df_proveedores = pd.DataFrame(datos_proveedor if isinstance(datos_proveedor, list) else [datos_proveedor])

    df_proveedores['VENDOR_ID'] = df_proveedores['VENDOR_ID'].astype(str).str.replace(r'\.0$', '',
                                                                                      regex=True).str.strip().str.upper()
    df_proveedores = df_proveedores.drop_duplicates(subset=['VENDOR_ID'])
    logging.info("Proveedores unicos listos para cruce: %d", len(df_proveedores))

    # =========================================================
    # 3. EXTRACCION FISCAL RECIENTE (SQL: CONT_Info_Fiscal filtrado por facturas I-)
    # =========================================================
    cols_fiscal = "MF_Nume1,MF_Alfa2,MF_Lista2,ID_Relacionado,MF_Bool5"

    logging.info("Fase 3/4: Descargando tabla fiscal filtrada por rango de facturas (I-)...")
    condicion_fiscal = "ID_Relacionado >= 'I-' AND ID_Relacionado < 'J-' AND MF_Bool5 = 0"

    while True:
        datos_fiscal = ejecutar_consulta_tabla(token_actual, cols_fiscal, "CONT_Info_Fiscal", False, condicion_fiscal, limite=250000)

        if datos_fiscal == "UNAUTHORIZED":
            logging.warning("La sesion expiro durante la extraccion masiva. Renovando token...")
            nuevo_token = obtener_token()
            if nuevo_token:
                token_actual = nuevo_token
                continue
            else:
                return None
        elif datos_fiscal == "ERROR" or not datos_fiscal:
            logging.error("Fallo critico en la recuperacion de la tabla secundaria.")
            return None
        else:
            break

    df_fiscal = pd.DataFrame(datos_fiscal)
    df_fiscal['ID_Relacionado'] = df_fiscal['ID_Relacionado'].astype(str).str.replace(r'\.0$', '',
                                                                                      regex=True).str.strip().str.upper()
    logging.info("Registros fiscales recuperados exitosamente: %d", len(df_fiscal))

    # =========================================================
    # 4. PROCESAMIENTO ATS (INNER JOINS EN RAM)
    # =========================================================
    logging.info("Fase 4/4: Consolidando reporte ATS en memoria RAM.")

    df_cruce = pd.merge(df_facturas, df_proveedores, on='VENDOR_ID', how='inner')
    df_cruce = pd.merge(df_cruce, df_fiscal, left_on='DOC_ID_CORP', right_on='ID_Relacionado', how='inner')

    if df_cruce.empty:
        logging.warning("Alerta ATS: El cruce relacional entre facturas e informacion fiscal arrojo 0 coincidencias.")
        return df_cruce

    df_cruce['ES_ANULADO'] = 0

    campos_numericos = ['TotalProductosConIVa', 'TotalServiciosConIVa', 'TotalProductosSinIVa', 'TotalServiciosSinIVa']
    for col in campos_numericos:
        df_cruce[col] = pd.to_numeric(df_cruce[col], errors='coerce').fillna(0)

    df_cruce['SUMA_CON_IVA'] = df_cruce['TotalProductosConIVa'] + df_cruce['TotalServiciosConIVa']
    df_cruce['SUMA_SIN_IVA'] = df_cruce['TotalProductosSinIVa'] + df_cruce['TotalServiciosSinIVa']

    df_cruce = df_cruce.sort_values(
        by=['MF_Lista2', 'INVOICE_DATE', 'DOC_REFERENCE'],
        ascending=[True, True, True]
    )

    df_cruce = df_cruce.head(100000)

    columnas_finales = [
        "INVOICE_DATE", "CORP", "VENDOR_ID", "VENDOR_NAME", "RUC_or_FED_ID",
        "MEMO", "INVOICE_TOTAL", "DOC_REFERENCE", "MF_Nume1", "MF_Alfa2",
        "MF_Lista2", "SUMA_CON_IVA", "SUMA_SIN_IVA", "ES_ANULADO",
        "DOC_ID_CORP", "ID_Relacionado"
    ]

    df_final = df_cruce[[col for col in columnas_finales if col in df_cruce.columns]]
    logging.info("Consolidacion ATS exitosa. Lineas netas preparadas para Excel: %d", len(df_final))
    return df_final


def exportar_a_excel_empresarial(df_datos, fecha_inicio, fecha_fin):
    if df_datos is None or df_datos.empty:
        logging.warning("Conjunto de datos vacio. Se omite la generacion del archivo Excel.")
        return

    logging.info("Exportando matriz a Excel corporativo con saneamiento de datos...")
    try:
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

        def sanitizar_excel(valor):
            if pd.isna(valor):
                return valor
            val_str = str(valor).strip()
            val_str = ILLEGAL_CHARACTERS_RE.sub("", val_str)
            if len(val_str) > 32700:
                return val_str[:32700] + "... [TRUNCADO por LIMITE]"
            return val_str

        for col in df_datos.select_dtypes(include=['object']).columns:
            df_datos[col] = df_datos[col].apply(sanitizar_excel)

        archivo_final = f"REPORTE_ATS_PERIODO_({fecha_inicio}_al_{fecha_fin}).xlsx"

        with pd.ExcelWriter(archivo_final, engine='openpyxl') as writer:
            df_datos.to_excel(writer, index=False, sheet_name='ATS_Consolidado')
            worksheet = writer.sheets['ATS_Consolidado']

            fill_encabezado = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            fuente_encabezado = Font(color="FFFFFF", bold=True)
            alineacion_centrada = Alignment(horizontal="center", vertical="center")

            for col_num in range(1, len(df_datos.columns) + 1):
                celda = worksheet.cell(row=1, column=col_num)
                celda.fill = fill_encabezado
                celda.font = fuente_encabezado
                celda.alignment = alineacion_centrada

            for idx, col_name in enumerate(df_datos.columns):
                ancho_calculado = max(
                    df_datos[col_name].astype(str).map(len).max(),
                    len(str(col_name))
                ) + 3
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(ancho_calculado, 35)

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

        logging.info("Archivo generado correctamente: %s", archivo_final)

    except Exception as e:
        logging.error("Fallo critico en creacion del archivo Excel: %s", e)


if __name__ == "__main__":
    fecha_inicio, fecha_fin = solicitar_rango_fechas()
    token_jwt = obtener_token()

    if token_jwt:
        df_resultado = procesar_cruce_ats(token_jwt, fecha_inicio, fecha_fin)
        exportar_a_excel_empresarial(df_resultado, fecha_inicio, fecha_fin)
    else:
        logging.critical("Ejecucion detenida. Servidor inaccesible.")