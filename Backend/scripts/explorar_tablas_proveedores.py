"""
Explora estructura+datos de tablas ERP (grupo PROVEEDORES) para diseñar reporte nuevo.
SELECT * LIMIT 40 por tabla, dump a Excel (una hoja por tabla).
Login con servicio pasado por env vars ERP_TEST_CODIGO/ERP_TEST_PASSWORD (no via settings,
para poder probar con un servicio distinto al configurado en .env).
"""
import sys
import os
import requests
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from app.config import settings
from app.repositories.mba3_repository import procesar_respuesta_erp

# tabla -> columna de fecha para filtrar (None = sin filtro, solo estructura)
TABLAS = {
    "SIST_Parametros_Empresa": None,
    "SIST_Parametros_Modulos": None,
}
LIMIT = 50
FECHA_DESDE = "2026-01-01"
FECHA_HASTA = "2026-01-31"

CODIGO = os.environ["ERP_TEST_CODIGO"]
PASSWORD = os.environ["ERP_TEST_PASSWORD"]
BASE_URL = settings.MBA3_BASE_URL_TEST

resp = requests.post(
    f"{BASE_URL}/ws2_mba3_serv_/login_servicio",
    json={"codigo": CODIGO, "pwd": PASSWORD},
    headers={"Content-Type": "application/json"},
    timeout=15,
)
resp.raise_for_status()
token = resp.json().get("jwt")
if not token:
    print(f"Login sin jwt en respuesta: {resp.json()}")
    sys.exit(1)
print(f"Login OK con servicio {CODIGO}, token obtenido.")

wb = Workbook()
wb.remove(wb.active)

for tabla, col_fecha in TABLAS.items():
    payload = {"select": "*", "from": tabla, "limit": str(LIMIT)}
    if col_fecha:
        payload["where"] = f"{col_fecha} >= '{FECHA_DESDE}' AND {col_fecha} <= '{FECHA_HASTA}'"
    print(f"Consultando {tabla} ({payload.get('where', 'sin filtro')})...")
    r = requests.post(
        f"{BASE_URL}/ws2_mba3_serv_Consultas_Externas_/",
        headers={"Authorization": token},
        data=payload,
        timeout=120,
    )
    r.raise_for_status()
    filas = procesar_respuesta_erp(r.json(), f"tabla {tabla}")
    ws = wb.create_sheet(title=tabla[:31])

    if not filas:
        ws["A1"] = f"Sin datos o error para {tabla}"
        print(f"  -> 0 filas")
        continue

    columnas = list(filas[0].keys())
    ws.append(columnas)
    for fila in filas:
        valores = []
        for c in columnas:
            v = fila.get(c)
            if isinstance(v, (dict, list)):
                v = str(v)
            valores.append(v)
        ws.append(valores)
    print(f"  -> {len(filas)} filas, {len(columnas)} columnas")

out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exploracion_proveedores.xlsx")
wb.save(out_path)
print(f"Excel guardado: {out_path}")
