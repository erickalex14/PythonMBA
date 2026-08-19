"""Carga inicial de las tablas KPI desde el Excel que hoy se arma a mano.

Llena `kpi_sucursal`, `kpi_producto_cat` y las metas del mes desde
"SEGUIMIENTO KPI AL ...xlsx". Es idempotente: se puede correr de nuevo.

Correr con:
    py -3 Backend/scripts/seed_kpi_desde_excel.py "C:/ruta/SEGUIMIENTO KPI.xlsx" 2026-08
"""
import re
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from app.core.database import SessionLocal  # noqa: E402

# Columna de la hoja RESUMEN KPI que trae la META de cada KPI (0-indexado).
COL_META = {
    4: "rentabilidad", 6: "tecnologia", 8: "celulares_tablets", 10: "motorola",
    12: "sillas_gamer", 14: "hogar_gym", 16: "planes_claro", 18: "review_env",
    20: "credito_directo", 22: "servicio_tecnico",
}

# "001 RIO COCA" -> codigo 001 + nombre. Descarta filas sueltas de la hoja
# ("CORTE AL ...", "SUCURSAL") que no son sucursales.
RE_SUCURSAL = re.compile(r"^\s*(\d{3})\s+(.+?)\s*$")


def sucursales_desde_presupuesto(wb):
    out = {}
    for fila in wb["PRESUPUESTO"].iter_rows(min_row=2, values_only=True):
        m = RE_SUCURSAL.match(str(fila[1] or ""))
        if not m:
            continue
        out[m.group(1)] = {
            "codigo": m.group(1), "nombre": m.group(2),
            "marca": fila[2], "ciudad": fila[3], "supervisor": fila[4],
        }
    return out


def catalogo_desde_pond(wb):
    """POND!J:L = codigo, producto, categoria. Es el VLOOKUP del Excel."""
    out = {}
    for fila in wb["POND"].iter_rows(min_row=2, min_col=10, max_col=12,
                                     values_only=True):
        codigo, producto, cat = fila
        if not codigo or not cat:
            continue
        out[str(codigo).strip().upper()] = (str(producto or "").strip(),
                                            str(cat).strip().upper())
    return out


def metas_desde_resumen(wb):
    out = []
    for fila in wb["RESUMEN KPI"].iter_rows(min_row=3, values_only=True):
        m = RE_SUCURSAL.match(str(fila[1] or ""))
        if not m:
            continue
        for col, kpi in COL_META.items():
            if col < len(fila) and isinstance(fila[col], (int, float)):
                out.append((m.group(1), kpi, float(fila[col])))
    return out


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        raise SystemExit(2)
    ruta, periodo = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    sucs = sucursales_desde_presupuesto(wb)
    cat = catalogo_desde_pond(wb)
    metas = metas_desde_resumen(wb)
    print(f"leidos: {len(sucs)} sucursales, {len(cat)} productos, {len(metas)} metas")

    db = SessionLocal()
    try:
        for s in sucs.values():
            db.execute(text("""
                INSERT INTO kpi_sucursal (codigo, nombre, supervisor, marca, ciudad, activa)
                VALUES (:codigo, :nombre, :supervisor, :marca, :ciudad, 'SI')
                ON CONFLICT (codigo) DO UPDATE SET
                    nombre = EXCLUDED.nombre, supervisor = EXCLUDED.supervisor,
                    marca = EXCLUDED.marca, ciudad = EXCLUDED.ciudad
            """), s)

        for codigo, (producto, c) in cat.items():
            db.execute(text("""
                INSERT INTO kpi_producto_cat (codigo, cat, producto)
                VALUES (:codigo, :cat, :producto)
                ON CONFLICT (codigo) DO UPDATE SET
                    cat = EXCLUDED.cat, producto = EXCLUDED.producto
            """), {"codigo": codigo, "cat": c, "producto": producto})

        for sucursal, kpi, meta in metas:
            db.execute(text("""
                INSERT INTO kpi_meta (periodo, sucursal, kpi, meta)
                VALUES (:p, :s, :k, :m)
                ON CONFLICT (periodo, sucursal, kpi) DO UPDATE SET meta = EXCLUDED.meta
            """), {"p": periodo, "s": sucursal, "k": kpi, "m": meta})

        db.commit()
        print(f"cargado en la base para el periodo {periodo}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
