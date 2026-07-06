import requests
import logging
import pandas as pd
import json
import datetime
import re
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configuracion del registro de eventos profesional
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Constantes de configuracion para produccion
BASE_URL = "http://181.198.104.181:8020"
URL_LOGIN = f"{BASE_URL}/ws2_mba3_serv_/login_servicio"
URL_CONSULTA = f"{BASE_URL}/ws2_mba3_serv_Consultas_Externas_/"

CODIGO_SERVICIO = "ERICKDEV"
PASSWORD_SERVICIO = "***REDACTED***"


def solicitar_fecha_reporte():
    print("\n--- Generador de Reporte Detallado de Ventas Espejo (MBA3) ---")
    while True:
        fecha_str = input("Ingrese la fecha a consultar (YYYY-MM-DD): ").strip()
        try:
            fecha_obj = datetime.datetime.strptime(fecha_str, "%Y-%m-%d")
            return fecha_obj
        except ValueError:
            print("Error: Formato incorrecto. Utilice YYYY-MM-DD (ejemplo: 2026-06-01).")


def obtener_token():
    logging.info("Solicitando credenciales de sesion al ERP...")
    headers = {"Content-Type": "application/json"}
    payload = {"codigo": CODIGO_SERVICIO, "pwd": PASSWORD_SERVICIO}

    try:
        response = requests.post(URL_LOGIN, json=payload, headers=headers, timeout=15)
        response.raise_for_status()
        datos = response.json()
        token = datos.get("jwt")
        if token:
            logging.info("Autenticacion exitosa. Token asignado.")
            return token
        return None
    except Exception as e:
        logging.error("Fallo de red al autenticar: %s", e)
        return None


def extraer_datos_kardex_detallado(token, fecha_consulta):
    headers = {"Authorization": token}
    fecha_str = fecha_consulta.strftime('%Y-%m-%d')
    condicion_where = f"TRANS_DATE = '{fecha_str}'"

    # Solicitamos las columnas del motor necesarias para calcular el reporte espejo detallado
    cols_movs = (
        "DOC_ID_CORP,TRANS_DATE,PRODUCT_ID_CORP,PRODUCT_NAME,QUANTITY,ORIGINAL_QTY,"
        "UNIT_COST,DISCOUNT_AMOUNT,NET_LINE_TOTAL,UM,Anulada,IN_OUT,"
        "\"Codigo grupo\",\"Codigo subgrupo\",Codigo_grupo,Codigo_subgrupo"
    )

    payload = {
        "select": cols_movs,
        "from": "INVT_Producto_Movimientos",
        "where": condicion_where,
        "limit": "200000"
    }

    try:
        logging.info("Descargando transacciones de inventario desde el ERP para el dia: %s...", fecha_str)
        response = requests.post(URL_CONSULTA, headers=headers, data=payload, timeout=180)
        response.raise_for_status()
        datos = response.json()

        if isinstance(datos, dict) and (datos.get("codigo") == "009" or datos.get("error") == "009"):
            logging.warning("No se encontraron transacciones registradas en la fecha solicitada.")
            return None

        df_bruto = pd.DataFrame(datos if isinstance(datos, list) else [datos])
        logging.info("Registros detallados extraidos: %d", len(df_bruto))
        return df_bruto

    except Exception as e:
        logging.error("Error en la extraccion de datos desde la API: %s", e)
        return None


def procesar_reporte_detallado(df_bruto):
    logging.info("Normalizando y estructurando columnas espejo en memoria RAM...")

    # 1. Normalizador Dinámico de Columnas contra las mañas de 4D
    columnas_reales = list(df_bruto.columns)
    mapeo_normalizado = {c.replace(" ", "").replace("_", "").upper(): c for c in columnas_reales}

    col_id_factura = mapeo_normalizado.get("DOCIDCORP")
    col_codigo_prod = mapeo_normalizado.get("PRODUCTIDCORP")
    col_nombre_prod = mapeo_normalizado.get("PRODUCTNAME")
    col_cant_real = mapeo_normalizado.get("ORIGINALQTY") if mapeo_normalizado.get(
        "ORIGINALQTY") else mapeo_normalizado.get("QUANTITY")
    col_grupo = mapeo_normalizado.get("CODIGOGRUPO")
    col_subgrupo = mapeo_normalizado.get("CODIGOSUBGRUPO")
    col_precio = mapeo_normalizado.get("UNITCOST")
    col_descuento = mapeo_normalizado.get("DISCOUNTAMOUNT")
    col_total_linea = mapeo_normalizado.get("NETLINETOTAL")

    # Renombramos a variables internas estables para operar de manera segura
    df_bruto = df_bruto.rename(columns={
        col_id_factura: 'FACTURA_INT',
        col_codigo_prod: 'CODIGO_INT',
        col_nombre_prod: 'PRODUCTO_INT',
        col_cant_real: 'CANTIDAD_INT',
        col_grupo: 'GRUPO_INT',
        col_subgrupo: 'SUBGRUPO_INT',
        col_precio: 'PRECIO_INT',
        col_descuento: 'DESCUENTO_INT',
        col_total_linea: 'TOTAL_INT'
    })

    # 2. Saneamiento de tipos de datos y strings
    df_bruto['FACTURA_INT'] = df_bruto['FACTURA_INT'].astype(str).str.replace(r'\.0$', '',
                                                                              regex=True).str.strip().str.upper()
    df_bruto['CODIGO_INT'] = df_bruto['CODIGO_INT'].astype(str).str.replace(r'\.0$', '',
                                                                            regex=True).str.strip().str.upper()
    df_bruto['PRODUCTO_INT'] = df_bruto['PRODUCTO_INT'].astype(str).str.strip().str.upper()
    df_bruto['UM'] = df_bruto['UM'].astype(str).str.strip().str.upper() if 'UM' in df_bruto.columns else 'UNID'
    df_bruto['IN_OUT'] = df_bruto['IN_OUT'].astype(str).str.strip().str.upper()

    # Castings numéricos precisos
    df_bruto['CANTIDAD_INT'] = pd.to_numeric(df_bruto['CANTIDAD_INT'], errors='coerce').fillna(0)
    df_bruto['PRECIO_INT'] = pd.to_numeric(df_bruto['PRECIO_INT'], errors='coerce').fillna(0.0)
    df_bruto['DESCUENTO_INT'] = pd.to_numeric(df_bruto['DESCUENTO_INT'], errors='coerce').fillna(0.0)
    df_bruto['TOTAL_INT'] = pd.to_numeric(df_bruto['TOTAL_INT'], errors='coerce').fillna(0.0)

    # Detección elástica de anulaciones
    def evaluar_anulada(val):
        if pd.isna(val): return False
        if isinstance(val, bool): return val
        return str(val).strip().lower() in ['true', '1', 't', 's', 'si', 'y']

    df_bruto['IS_ANULADA'] = df_bruto['Anulada'].apply(evaluar_anulada) if 'Anulada' in df_bruto.columns else False

    # 3. Filtros Ejecutivos (Solo ventas legítimas no anuladas con cantidades físicas reales)
    df_filtrado = df_bruto[
        (df_bruto['IS_ANULADA'] == False) &
        (df_bruto['CANTIDAD_INT'] > 0)
        ].copy()

    if df_filtrado.empty:
        logging.warning("No existen registros de transacciones que cumplan las condiciones de venta.")
        return None

    # 4. Cálculos dinámicos e inyección estructural de las columnas del MBA3
    df_filtrado['CANTIDAD_INT'] = df_filtrado['CANTIDAD_INT'].round(0).astype(int)
    df_filtrado['SUBTOTAL_INT'] = df_filtrado['CANTIDAD_INT'] * df_filtrado['PRECIO_INT']

    # Si por mañas de 4D el TOTAL LINEA viene en cero, lo recalculamos automáticamente
    df_filtrado['TOTAL_INT'] = df_filtrado.apply(
        lambda r: r['TOTAL_INT'] if r['TOTAL_INT'] > 0 else (r['SUBTOTAL_INT'] - r['DESCUENTO_INT']), axis=1
    )

    # Estructuramos la salida con los nombres exactos y el orden riguroso que pediste
    df_espejo = pd.DataFrame()
    df_espejo['# de factura'] = df_filtrado['FACTURA_INT']
    df_espejo['FECHA'] = df_filtrado['TRANS_DATE']
    df_espejo['CODIGO'] = df_filtrado['CODIGO_INT']
    df_espejo['PRODUCTO'] = df_filtrado['PRODUCTO_INT']
    df_espejo['GRUPO'] = df_filtrado['GRUPO_INT'].fillna('GENERAL')
    df_espejo['SUBGRUPO'] = df_filtrado['SUBGRUPO_INT'].fillna('GENERAL')
    df_espejo['UNIDAD'] = df_filtrado['UM']
    df_espejo['CANTIDAD'] = df_filtrado['CANTIDAD_INT']
    df_espejo['PRECIO VENTA'] = df_filtrado['PRECIO_INT'].round(4)
    df_espejo['SUBTOTAL (C*PV)'] = df_filtrado['SUBTOTAL_INT'].round(4)
    df_espejo['DESCUENTO APLICADO'] = df_filtrado['DESCUENTO_INT'].round(4)
    df_espejo['TOTAL LINEA'] = df_filtrado['TOTAL_INT'].round(4)

    # Ordenamos cronológicamente por factura y código para mantener la consistencia corporativa
    df_espejo = df_espejo.sort_values(by=['# de factura', 'CODIGO'], ascending=[True, True])
    return df_espejo


def exportar_a_excel_empresarial(df_datos, fecha_consulta):
    if df_datos is None or df_datos.empty:
        return

    logging.info("Generando renderizado final en archivo Excel estructurado...")
    try:
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

        def sanitizar_excel(valor):
            if pd.isna(valor): return valor
            val_str = str(valor).strip()
            val_str = ILLEGAL_CHARACTERS_RE.sub("", val_str)
            return val_str

        for col in df_datos.select_dtypes(include=['object']).columns:
            df_datos[col] = df_datos[col].apply(sanitizar_excel)

        fecha_str = fecha_consulta.strftime('%Y-%m-%d')
        archivo_final = f"DETALLE_VENTAS_ESPEJO_({fecha_str}).xlsx"

        with pd.ExcelWriter(archivo_final, engine='openpyxl') as writer:
            df_datos.to_excel(writer, index=False, sheet_name='Detalle Productos - Periodo')
            worksheet = writer.sheets['Detalle Productos - Periodo']

            fill_encabezado = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            fuente_encabezado = Font(color="FFFFFF", bold=True)
            alineacion_centrada = Alignment(horizontal="center", vertical="center")

            # Aplicar formato de cabecera corporativa MBA
            for col_num in range(1, len(df_datos.columns) + 1):
                celda = worksheet.cell(row=1, column=col_num)
                celda.fill = fill_encabezado
                celda.font = fuente_encabezado
                celda.alignment = alineacion_centrada

            # Ancho de columnas adaptable automático
            for idx, col_name in enumerate(df_datos.columns):
                ancho_calculado = max(
                    df_datos[col_name].astype(str).map(len).max(),
                    len(str(col_name))
                ) + 4
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(ancho_calculado, 50)

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

        logging.info("¡GOLAZO LOCO! Reporte transaccional generado idéntico al ERP: %s", archivo_final)

    except Exception as e:
        logging.error("Fallo critico I/O al escribir el archivo Excel: %s", e)


if __name__ == "__main__":
    fecha_reporte = solicitar_fecha_reporte()
    token_jwt = obtener_token()

    if token_jwt:
        df_bruto = extraer_datos_kardex_detallado(token_jwt, fecha_reporte)
        if df_bruto is not None:
            df_final = procesar_reporte_detallado(df_bruto)
            exportar_a_excel_empresarial(df_final, fecha_reporte)
    else:
        logging.critical("Ejecucion abortada por falta de credenciales autorizadas.")