import io
import re
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')


def _clean(v):
    if pd.isna(v):
        return ""
    s = ILLEGAL_CHARACTERS_RE.sub("", str(v).strip())
    return s[:32700]


class ExcelService:
    """
    Servicio de Utilidad para la Generación de Archivos Excel.
    Todos los reportes (Ventas, Movimientos, Liquidaciones, ATS) comparten el mismo
    formato corporativo: encabezado, resumen de totales, tabla con estilo y fila de TOTALES.
    """

    def generar_reporte_excel(self, df: pd.DataFrame, sheet_name: str) -> io.BytesIO:
        """
        Fallback genérico (sin encabezado/resumen/totales) para datos ad-hoc que no
        tienen un generador corporativo dedicado.
        """
        df_datos = df.copy()
        for col in df_datos.select_dtypes(include=['object']).columns:
            df_datos[col] = df_datos[col].apply(_clean)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_datos.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]

            fill_encabezado = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            fuente_encabezado = Font(color="FFFFFF", bold=True)
            alineacion_centrada = Alignment(horizontal="center", vertical="center")

            for col_num in range(1, len(df_datos.columns) + 1):
                celda = worksheet.cell(row=1, column=col_num)
                celda.fill = fill_encabezado
                celda.font = fuente_encabezado
                celda.alignment = alineacion_centrada

            for idx, col_name in enumerate(df_datos.columns):
                ancho_calculado = max(
                    df_datos[col_name].astype(str).map(len).max(),
                    len(str(col_name))
                ) + 3
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(ancho_calculado, 40)

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

        output.seek(0)
        return output

    def _generar_reporte_corporativo(
        self, df: pd.DataFrame, sheet_name: str, titulo: str, inicio: str, fin: str,
        columnas: list, money_cols: set, qty_cols: set = None, bool_cols: set = None,
        resumen: list = None, anchos: dict = None
    ) -> io.BytesIO:
        """
        Renderer compartido de una sola hoja: crea su propio workbook y lo devuelve.
        """
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if "Sheet" in writer.book.sheetnames:
                del writer.book["Sheet"]
            self._escribir_hoja_corporativa(
                writer.book, df, sheet_name, titulo, inicio, fin,
                columnas, money_cols, qty_cols, bool_cols, resumen, anchos
            )
        output.seek(0)
        return output

    def _escribir_hoja_corporativa(
        self, wb, df: pd.DataFrame, sheet_name: str, titulo: str, inicio: str, fin: str,
        columnas: list, money_cols: set, qty_cols: set = None, bool_cols: set = None,
        resumen: list = None, anchos: dict = None
    ) -> None:
        """
        Escribe UNA hoja con formato corporativo (encabezado, resumen, tabla, totales)
        dentro de un workbook ya existente - permite varias hojas en un mismo archivo.
        """
        qty_cols = qty_cols or set()
        bool_cols = bool_cols or set()
        anchos = anchos or {}
        df = df.copy()

        cols_presentes = [(src, lbl) for src, lbl in columnas if src in df.columns]

        def num(col):
            return pd.to_numeric(df[col], errors='coerce').fillna(0) if col in df.columns else pd.Series([0] * len(df))

        if True:
            ws = wb.create_sheet(sheet_name)

            azul = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            gris = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            verde = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            blanco_bold = Font(color="FFFFFF", bold=True)
            bold = Font(bold=True)
            centro = Alignment(horizontal="center", vertical="center")
            borde = Border(bottom=Side(style="thin", color="B0B0B0"))

            # --- Encabezado corporativo ---
            ws.cell(row=1, column=1, value="NOVISOLUTIONS CIA. LTDA.").font = Font(bold=True, size=14, color="002060")
            ws.cell(row=2, column=1, value=titulo).font = Font(bold=True, size=11)
            ws.cell(row=3, column=1, value=f"Rango: {inicio}  al  {fin}").font = Font(italic=True, color="555555")

            # --- Resumen (opcional) ---
            r = 5
            for etq, val in (resumen or []):
                c1 = ws.cell(row=r, column=1, value=etq)
                c1.font = bold
                c1.fill = verde
                c2 = ws.cell(row=r, column=2, value=round(val, 2) if isinstance(val, float) else val)
                c2.font = bold
                c2.fill = verde
                if isinstance(val, float):
                    c2.number_format = '#,##0.00'
                r += 1

            # --- Cabecera de la tabla ---
            hdr_row = r + 1
            for j, (src, lbl) in enumerate(cols_presentes, start=1):
                cel = ws.cell(row=hdr_row, column=j, value=lbl)
                cel.fill = azul
                cel.font = blanco_bold
                cel.alignment = centro

            # --- Filas de detalle ---
            data_start = hdr_row + 1
            rr = data_start
            for _, fila in df.iterrows():
                for j, (src, lbl) in enumerate(cols_presentes, start=1):
                    val = fila.get(src)
                    if src in money_cols or src in qty_cols:
                        try:
                            val = float(val)
                        except Exception:
                            val = 0
                        cel = ws.cell(row=rr, column=j, value=val)
                        cel.number_format = '#,##0.00' if src in money_cols else '#,##0'
                    elif src in bool_cols:
                        cel = ws.cell(row=rr, column=j, value="SI" if _to_bool(val) else "NO")
                    else:
                        cel = ws.cell(row=rr, column=j, value=_clean(val))
                    cel.border = borde
                rr += 1

            # --- Fila de TOTALES ---
            tot_row = rr
            ws.cell(row=tot_row, column=1, value=f"TOTALES - {len(df)} líneas").font = bold
            for j, (src, lbl) in enumerate(cols_presentes, start=1):
                if src in qty_cols:
                    c = ws.cell(row=tot_row, column=j, value=float(num(src).sum()))
                    c.number_format = '#,##0'; c.font = bold; c.fill = gris
                elif src in money_cols:
                    c = ws.cell(row=tot_row, column=j, value=round(float(num(src).sum()), 2))
                    c.number_format = '#,##0.00'; c.font = bold; c.fill = gris
                else:
                    ws.cell(row=tot_row, column=j).fill = gris

            # Anchos de columna
            for j, (src, lbl) in enumerate(cols_presentes, start=1):
                ws.column_dimensions[get_column_letter(j)].width = anchos.get(src, 14)

            ws.freeze_panes = ws.cell(row=data_start, column=1)

    def generar_reporte_ventas(self, df: pd.DataFrame, inicio: str, fin: str) -> io.BytesIO:
        df = df.copy()
        # Normalizar claves: el GET manda "# de factura"/"TOTAL LINEA"; el POST
        # (datos filtrados del front) manda factura_final/total_linea. Unificar.
        rename_dto = {
            "factura_final": "# de factura", "fecha": "FECHA", "empresa": "EMPRESA",
            "sucursal": "SUCURSAL", "codigo": "CODIGO", "producto": "PRODUCTO",
            "grupo": "GRUPO", "subgrupo": "SUBGRUPO", "unidad": "UNIDAD", "cantidad": "CANTIDAD",
            "precio_venta": "PRECIO VENTA", "subtotal": "SUBTOTAL (C*PV)",
            "descuento_aplicado": "DESCUENTO APLICADO", "total_linea": "TOTAL LINEA",
            "bodega": "BODEGA", "bodega_nombre": "BODEGA NOMBRE", "codigo_cliente": "CODIGO CLIENTE",
            "nombre_cliente": "NOMBRE CLIENTE", "costo_unitario": "COSTO UNITARIO",
            "costo_total": "COSTO TOTAL", "utilidad_unidad": "UTILIDAD UNIDAD",
            "utilidad_total": "UTILIDAD TOTAL", "pct_utilidad_neto": "% UTILIDAD/NETO",
            "pct_utilidad_costo": "% UTILIDAD/COSTO",
        }
        df = df.rename(columns={k: v for k, v in rename_dto.items() if k in df.columns})

        columnas = [
            ("# de factura", "# de factura"), ("FECHA", "Fecha"), ("EMPRESA", "Empresa"),
            ("SUCURSAL", "Sucursal"), ("CODIGO", "Código"), ("PRODUCTO", "Producto"),
            ("GRUPO", "Grupo"), ("SUBGRUPO", "Subgrupo"), ("UNIDAD", "Unidad"),
            ("CANTIDAD", "Cantidad"), ("PRECIO VENTA", "Precio Venta"),
            ("SUBTOTAL (C*PV)", "SubTotal (C*PV)"), ("DESCUENTO APLICADO", "Descuento Aplicado"),
            ("TOTAL LINEA", "Total Linea"), ("BODEGA", "Bodega"), ("BODEGA NOMBRE", "Bodega Nombre"),
            ("CODIGO CLIENTE", "Código Cliente"), ("NOMBRE CLIENTE", "Nombre Cliente"),
            ("COSTO UNITARIO", "Costo Unitario"), ("COSTO TOTAL", "Costo Total"),
            ("UTILIDAD UNIDAD", "Utilidad Unidad"), ("UTILIDAD TOTAL", "Utilidad Total"),
            ("% UTILIDAD/NETO", "% Utilidad/Neto"), ("% UTILIDAD/COSTO", "% Utilidad/Costo"),
        ]
        money_cols = {"PRECIO VENTA", "SUBTOTAL (C*PV)", "DESCUENTO APLICADO", "TOTAL LINEA",
                      "COSTO UNITARIO", "COSTO TOTAL", "UTILIDAD UNIDAD", "UTILIDAD TOTAL",
                      "% UTILIDAD/NETO", "% UTILIDAD/COSTO"}

        def num(col):
            return pd.to_numeric(df[col], errors='coerce').fillna(0) if col in df.columns else pd.Series([0] * len(df))
        total_general = float(num("TOTAL LINEA").sum())
        emp_series = df["EMPRESA"].astype(str) if "EMPRESA" in df.columns else pd.Series([""] * len(df))
        suc_series = df["SUCURSAL"].astype(str) if "SUCURSAL" in df.columns else pd.Series([""] * len(df))
        tl = num("TOTAL LINEA")
        resumen = [
            ("TOTAL VENTAS (ENV + Novicompu)", total_general),
            ("Novicompu", float(tl[emp_series.str.upper().str.contains("NOVI")].sum())),
            ("ENV", float(tl[emp_series.str.upper().str.contains("ENV")].sum())),
            ("Mayoristas (suc. 026 + 027)", float(tl[suc_series.isin(["026", "027"])].sum())),
            ("Unidades vendidas", float(num("CANTIDAD").sum())),
        ]
        anchos = {"# de factura": 18, "PRODUCTO": 38, "CODIGO": 16, "GRUPO": 10, "SUBGRUPO": 10,
                  "EMPRESA": 12, "SUCURSAL": 10, "FECHA": 12, "BODEGA NOMBRE": 22, "NOMBRE CLIENTE": 28}

        return self._generar_reporte_corporativo(
            df, "Detalle Productos - Periodo", "Reporte de Ventas - Detalle Productos por Período",
            inicio, fin, columnas, money_cols, qty_cols={"CANTIDAD"}, resumen=resumen, anchos=anchos
        )

    def generar_reporte_movimientos(self, df: pd.DataFrame, inicio: str, fin: str) -> io.BytesIO:
        columnas = [
            ("TRANS_DATE", "Fecha"), ("Codigo_producto_convertido", "Código"),
            ("PRODUCT_NAME", "Producto"), ("ORIGINAL_QTY", "Cantidad"),
            ("ORIGIN_MEMO", "Origen"), ("ORIGIN_REF", "Referencia"),
            ("Codigo_Sucursal", "Sucursal"), ("Codigo_Marca", "Marca"),
            ("COD_SALESMAN", "Vendedor"), ("BASE_COMISION", "Base Comisión"),
            ("BaseImponibleReal_1", "Base Imponible"), ("Info_Seriales", "Seriales"),
        ]
        money_cols = {"BASE_COMISION", "BaseImponibleReal_1"}

        def num(col):
            return pd.to_numeric(df[col], errors='coerce').fillna(0) if col in df.columns else pd.Series([0] * len(df))
        resumen = [
            ("Total Movimientos", len(df)),
            ("Total Cantidad", float(num("ORIGINAL_QTY").sum())),
            ("Total Base Comisión", float(num("BASE_COMISION").sum())),
            ("Total Base Imponible", float(num("BaseImponibleReal_1").sum())),
        ]
        anchos = {"PRODUCT_NAME": 38, "Codigo_producto_convertido": 16, "Info_Seriales": 30, "TRANS_DATE": 12}

        return self._generar_reporte_corporativo(
            df, "Movimientos", "Reporte de Movimientos (Kardex/Seriales)",
            inicio, fin, columnas, money_cols, qty_cols={"ORIGINAL_QTY"}, resumen=resumen, anchos=anchos
        )

    def generar_reporte_liquidaciones(self, df: pd.DataFrame, inicio: str, fin: str) -> io.BytesIO:
        columnas = [
            ("CORP", "Corp"), ("LIQUIDACION_FECHA", "Fecha"), ("LIQUIDACION_ID_CORP", "Liquidación"),
            ("LIQUIDACION_ESTADO", "Estado"), ("FACTURA_ID_CORP", "Factura"),
            ("PRODUCTO_ID_CORP", "Producto"), ("CANTIDAD", "Cantidad"), ("PRECIO", "Precio"),
            ("TOTAL", "Total"), ("VALOR_TOTAL_CIF", "Valor Total CIF"),
            ("VALOR_SUBTOTAL_CIF", "Valor Subtotal CIF"), ("VALOR_TOTAL_CIF_MANUAL", "Valor CIF Manual"),
            ("VALOR_TOTAL_CIF_UNIDAD", "Valor CIF Unidad"),
            ("ANTES_TOTAL_1", "Antes Total 1"), ("ANTES_TOTAL_2", "Antes Total 2"), ("ANTES_TOTAL_3", "Antes Total 3"),
            ("DESPUES_TOTAL_1", "Después Total 1"), ("DESPUES_TOTAL_2", "Después Total 2"), ("DESPUES_TOTAL_3", "Después Total 3"),
            ("VALOR_ANTES_1", "Valor Antes 1"), ("VALOR_ANTES_2", "Valor Antes 2"), ("VALOR_ANTES_3", "Valor Antes 3"),
            ("VALOR_DESPUES_1", "Valor Después 1"), ("VALOR_DESPUES_2", "Valor Después 2"), ("VALOR_DESPUES_3", "Valor Después 3"),
            ("OBSERVACIONES", "Observaciones"), ("PARTIDA_ID_CORP", "Partida"), ("IdRecepcionRelacionada", "Recepción Relacionada"),
        ]
        money_cols = {"PRECIO", "TOTAL", "VALOR_TOTAL_CIF", "VALOR_SUBTOTAL_CIF", "VALOR_TOTAL_CIF_MANUAL",
                      "VALOR_TOTAL_CIF_UNIDAD", "ANTES_TOTAL_1", "ANTES_TOTAL_2", "ANTES_TOTAL_3",
                      "DESPUES_TOTAL_1", "DESPUES_TOTAL_2", "DESPUES_TOTAL_3",
                      "VALOR_ANTES_1", "VALOR_ANTES_2", "VALOR_ANTES_3",
                      "VALOR_DESPUES_1", "VALOR_DESPUES_2", "VALOR_DESPUES_3"}

        def num(col):
            return pd.to_numeric(df[col], errors='coerce').fillna(0) if col in df.columns else pd.Series([0] * len(df))
        resumen = [
            ("Total Liquidaciones (líneas)", len(df)),
            ("Total Valor CIF", float(num("VALOR_TOTAL_CIF").sum())),
            ("Total General", float(num("TOTAL").sum())),
        ]
        anchos = {"OBSERVACIONES": 30, "FACTURA_ID_CORP": 18, "PRODUCTO_ID_CORP": 18, "LIQUIDACION_ID_CORP": 18}

        return self._generar_reporte_corporativo(
            df, "Consolidado", "Reporte Consolidado de Liquidaciones",
            inicio, fin, columnas, money_cols, qty_cols={"CANTIDAD"}, resumen=resumen, anchos=anchos
        )

    def generar_reporte_estadisticas_ventas(self, df: pd.DataFrame, inicio: str, fin: str) -> io.BytesIO:
        """
        Replica el reporte nativo "Estadisticas de Inventarios" del ERP: hoja principal
        (todos los productos) + 2 hojas de Top 10 (por unidades y por dólares vendidos).
        """
        columnas = [
            ("codigo", "Código"), ("producto", "Descripción"), ("unidad", "Unidad"),
            ("grupo", "Grupo"), ("subgrupo", "Subgrupo"),
            ("existencia", "Existencia"), ("asignado", "Asignado"), ("disponible", "Disponible"),
            ("unidades_vendidas", "Unidades Vendidas"), ("total_ventas", "Total Ventas"),
            ("precio_promedio", "Precio Promedio"), ("precio_maximo", "Precio Máximo"),
            ("precio_minimo", "Precio Mínimo"), ("ultimo_precio", "Último Precio"),
            ("ultima_fecha_venta", "Última Fecha Venta"), ("no_dias", "No. Días"),
        ]
        money_cols = {"total_ventas", "precio_promedio", "precio_maximo", "precio_minimo", "ultimo_precio"}
        qty_cols = {"existencia", "asignado", "disponible", "unidades_vendidas", "no_dias"}
        anchos = {"producto": 34, "codigo": 16}

        def num(col):
            return pd.to_numeric(df[col], errors='coerce').fillna(0) if col in df.columns else pd.Series([0] * len(df))
        resumen = [
            ("Total Productos", len(df)),
            ("Unidades Vendidas", float(num("unidades_vendidas").sum())),
            ("Total Ventas", float(num("total_ventas").sum())),
        ]

        top_cantidad = df.sort_values(by="unidades_vendidas", ascending=False).head(10) if "unidades_vendidas" in df.columns else df.head(0)
        top_dolares = df.sort_values(by="total_ventas", ascending=False).head(10) if "total_ventas" in df.columns else df.head(0)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if "Sheet" in writer.book.sheetnames:
                del writer.book["Sheet"]
            self._escribir_hoja_corporativa(
                writer.book, df, "Estadisticas de Inventarios", "Reporte de Ventas - Estadísticas por Producto",
                inicio, fin, columnas, money_cols, qty_cols=qty_cols, resumen=resumen, anchos=anchos
            )
            self._escribir_hoja_corporativa(
                writer.book, top_cantidad, "Mas Vendido en Cantidades", "Top 10 Productos por Unidades Vendidas",
                inicio, fin, columnas, money_cols, qty_cols=qty_cols, anchos=anchos
            )
            self._escribir_hoja_corporativa(
                writer.book, top_dolares, "Mas Vendido en Dolares", "Top 10 Productos por Total Vendido ($)",
                inicio, fin, columnas, money_cols, qty_cols=qty_cols, anchos=anchos
            )
        output.seek(0)
        return output

    def generar_reporte_ats(self, df: pd.DataFrame, inicio: str, fin: str) -> io.BytesIO:
        columnas = [
            ("CORP", "Corp"), ("INVOICE_DATE", "Fecha"), ("VENDOR_ID", "Cód. Proveedor"),
            ("VENDOR_NAME", "Proveedor"), ("RUC_or_FED_ID", "RUC"), ("DOC_REFERENCE", "Referencia"),
            ("MEMO", "Concepto"), ("INVOICE_TOTAL", "Total Factura"),
            ("TotalProductosConIVa", "Productos Con IVA"), ("TotalServiciosConIVa", "Servicios Con IVA"),
            ("TotalProductosSinIVa", "Productos Sin IVA"), ("TotalServiciosSinIVa", "Servicios Sin IVA"),
            ("SUMA_CON_IVA", "Suma Con IVA"), ("SUMA_SIN_IVA", "Suma Sin IVA"),
            ("MF_Alfa2", "Fiscal Alfa 2"), ("MF_Alfa3", "Fiscal Alfa 3"), ("MF_Lista2", "Lista Fiscal"),
            ("MF_Nume1", "Fiscal Numérico 1"), ("Reservado1", "Reservado 1"),
            ("Reservado2", "Reservado 2"), ("Reservado3", "Reservado 3"),
            ("CONFIRMED", "Confirmada"), ("VOID", "Anulada"),
        ]
        money_cols = {"INVOICE_TOTAL", "TotalProductosConIVa", "TotalServiciosConIVa",
                      "TotalProductosSinIVa", "TotalServiciosSinIVa", "SUMA_CON_IVA", "SUMA_SIN_IVA", "MF_Nume1"}
        bool_cols = {"Reservado1", "Reservado2", "Reservado3", "CONFIRMED", "VOID"}

        def num(col):
            return pd.to_numeric(df[col], errors='coerce').fillna(0) if col in df.columns else pd.Series([0] * len(df))
        confirmadas = int(df["CONFIRMED"].apply(_to_bool).sum()) if "CONFIRMED" in df.columns else 0
        anuladas = int(df["VOID"].apply(_to_bool).sum()) if "VOID" in df.columns else 0
        resumen = [
            ("Total Facturado", float(num("INVOICE_TOTAL").sum())),
            ("Total Con IVA", float(num("SUMA_CON_IVA").sum())),
            ("Total Sin IVA", float(num("SUMA_SIN_IVA").sum())),
            ("Facturas Confirmadas", confirmadas),
            ("Facturas Anuladas", anuladas),
        ]
        anchos = {"VENDOR_NAME": 30, "MEMO": 26, "DOC_REFERENCE": 16}

        return self._generar_reporte_corporativo(
            df, "Consolidado", "Reporte de Facturación Fiscal (ATS)",
            inicio, fin, columnas, money_cols, bool_cols=bool_cols, resumen=resumen, anchos=anchos
        )


def _to_bool(val):
    if pd.isna(val):
        return False
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() in ['true', '1', 't', 'y', 'yes', 's', 'si']
