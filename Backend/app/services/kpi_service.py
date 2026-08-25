import calendar
import datetime
import io
import logging
import re
from typing import Optional

import openpyxl
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.database import SessionLocal

# --- Importacion del Excel que hoy se arma a mano -------------------------
# Columna de la hoja RESUMEN KPI que trae la META de cada KPI (0-indexado).
COL_META_EXCEL = {
    4: "rentabilidad", 6: "tecnologia", 8: "celulares_tablets", 10: "motorola",
    12: "sillas_gamer", 14: "hogar_gym", 16: "planes_claro", 18: "review_env",
    20: "credito_directo", 22: "servicio_tecnico",
}

# "001 RIO COCA" -> codigo + nombre. Descarta las filas sueltas de la hoja
# ("CORTE AL ...", "SUCURSAL") que no son sucursales.
RE_SUCURSAL_EXCEL = re.compile(r"^\s*(\d{3})\s+(.+?)\s*$")

# Definicion de los KPIs del reporte de Seguimiento.
#
# `peso` sale de la fila 2 de la hoja RESUMEN KPI del Excel, que es la que de
# verdad entra en las formulas de cumplimiento. La hoja POND tiene su propia
# columna de ponderacion y NO coincide en 6 de los 10 KPIs (ambas suman 30%, por
# eso la diferencia no salta al cuadrar el total). Pendiente de confirmar cual es
# la oficial; mientras tanto manda la que se usaba para calcular.
#
# `origen`:
#   ventas  -> se agrega desde view_ventas_espejo_reporte cruzando el catalogo
#   manual  -> lo captura una persona (hoy llega por un Google Form)
#   externo -> lo entrega otro proceso y este servicio solo lo recibe
KPIS = {
    # (venta - costo) / venta sobre TODA la venta de la tienda, no solo las
    # categorias que puntuan: verificado contra el archivo manual (002 da
    # 0.273749 contra 0.273751, 021 da 0.473613 contra 0.473622).
    "rentabilidad": {
        "label": "RENTABILIDAD DE TIENDA", "peso": 0.04,
        "origen": "margen", "medida": "ratio"},
    "tecnologia": {
        "label": "TECNOLOGIA / PROYECTORES", "peso": 0.04,
        "origen": "ventas", "cat": "TECNOLOGIA", "medida": "unidades"},
    "celulares_tablets": {
        "label": "CELULARES Y TABLETS ENV", "peso": 0.03,
        "origen": "ventas", "cat": "CELULARES Y TABLETS ENV", "medida": "unidades"},
    "motorola": {
        "label": "MOTOROLA RAZR / EDGE 60", "peso": 0.00,
        "origen": "ventas", "cat": "MOTOROLA RAZR/EDGE 60", "medida": "unidades"},
    "sillas_gamer": {
        "label": "SILLAS GAMER ENV", "peso": 0.02,
        "origen": "ventas", "cat": "SILLAS GAMER ENV", "medida": "unidades"},
    "hogar_gym": {
        "label": "HOGAR Y GIMNASIO", "peso": 0.02,
        "origen": "ventas", "cat": "HOGAR Y GYM", "medida": "unidades"},
    "servicio_tecnico": {
        "label": "SERVICIO TECNICO", "peso": 0.02,
        "origen": "ventas", "cat": "ST", "medida": "monto"},
    "planes_claro": {
        "label": "PLANES CLARO", "peso": 0.00,
        "origen": "manual", "medida": "unidades"},
    "review_env": {
        "label": "REVIEW ENV", "peso": 0.03,
        "origen": "manual", "medida": "unidades"},
    # Sale de los cobros del ERP: pagos 'Otros' de CrediNovi o Banco Solidario.
    "credito_directo": {
        "label": "CREDITO DIRECTO", "peso": 0.10,
        "origen": "cobros", "medida": "monto"},
}

CATS_VENTAS = {k: v["cat"] for k, v in KPIS.items() if v["origen"] == "ventas"}

# Meta de venta total de la tienda. No es un KPI ponderado: es el presupuesto
# mensual, y va en `kpi_meta` con esta llave para no inventarle otra tabla.
KPI_VENTA_TIENDA = "venta_tienda"
METAS_EXTRA = {KPI_VENTA_TIENDA: "META DE TIENDA"}

# Nombre de la hoja de detalle por categoria en el Excel que se arma a mano.
HOJA_POR_KPI = {
    "tecnologia": "TCNLG", "celulares_tablets": "TAB-CEL",
    "motorola": "RAZR-EDGE60", "sillas_gamer": "S.GAMER",
    "hogar_gym": "HOGAR-GYM", "servicio_tecnico": "ST",
}

# Encabezados de la hoja RESUMEN KPI, copiados tal cual del archivo que se arma
# a mano (con sus espacios de mas y su acentuacion): el reporte generado tiene
# que poder pegarse encima del anterior sin que nadie note la diferencia.
# (kpi, titulo de la columna del valor real, titulo de la columna de la meta)
COLUMNAS_RESUMEN = [
    ("rentabilidad", "RENTABILIDAD OBTENIDA", "RENTABILIDAD DE TIENDA"),
    ("tecnologia", "VENTA  DE TECNOLOGIA ", "TECNOLOGIA"),
    ("celulares_tablets", "VENTA CELULARES Y TABLETS ENV", "CELULARES Y TABLETS ENV"),
    ("motorola", "VTA MOTOROLA RAZER / EDGE 60", "MOTOROLA RAZER / EDGE 60"),
    ("sillas_gamer", "VENTA SILLAS GAMER ENV ", "SILLAS GAMER ENV "),
    ("hogar_gym", "VENTA HOGAR Y GYM", "HOGAR Y GIMNASIO"),
    ("planes_claro", "VTA PLANES CLARO", "PLANES CLARO"),
    ("review_env", "REVIEW ENV REALIZADO", "REVIEW ENV"),
    ("credito_directo", "Venta credito", "CREDITO DIRECTO"),
    ("servicio_tecnico", "VENTA SERVICIO TECNICO", "SERVICIO TECNICO"),
]

# Titulos del bloque de aportes (columnas Y..AH). No coinciden con los de arriba:
# el archivo original abrevia "HOGAR" y repite "VTA" en dos de ellos.
TITULOS_APORTE = {
    "rentabilidad": "RENTABILIDAD DE TIENDA", "tecnologia": "TECNOLOGIA",
    "celulares_tablets": "CELULARES Y TABLETS ENV",
    "motorola": "VTA MOTOROLA RAZER / EDGE 60", "sillas_gamer": "SILLAS GAMER ENV ",
    "hogar_gym": "HOGAR", "planes_claro": "VTA PLANES CLARO",
    "review_env": "REVIEW ENV", "credito_directo": "CREDITO DIRECTO",
    "servicio_tecnico": "SERVICIO TECNICO",
}

COLUMNAS_PRESUPUESTO = [
    "l", "SUCURSAL", "MARCA", "CIUDAD", "SUPERVISOR", "META DE TIENDA AGOSTO",
    "VENTA TOTAL DE TIENDA", "TOTAL DE FACTURAS TIENDA ",
    "TOTAL UNIDADES VENDIDAS X TIENDA", "Ticket promedio x factura (TIENDA)",
    "Unidades Promedio x factura UPF (TIENDA)", "OBSERVACION",
    "VTA PROMEDIO POR DIA", "PROYECCCION A FIN DE MES ", "CUMPLIMIENTO",
    "RENT", "KPI",
]

COLUMNAS_DETALLE = [
    ("factura_final", "No. Factura"), ("codigo_vendedor", "Código"),
    ("nombre_cliente", "Nombre"), ("canal", "CANAL"), ("fecha", "Fecha"),
    ("bodega_codigo", "Bodega"), ("sucursal_larga", "SUCURSAL"),
    ("supervisor", "SUPERVISOR"), ("codigo", "Código"), ("producto", "Producto"),
    ("unidad", "Unidad"), ("grupo", "Grupo"), ("subgrupo", "Subgrupo"),
    ("cantidad", "Cantidad"), ("precio_venta", "Precio de Venta"),
    ("total_linea", "Total Factura"), ("cat", "CAT"),
]


def observacion_tienda(ticket: float, upf: float) -> str:
    """Etiqueta de la columna OBSERVACION, con los cortes del Excel original."""
    if ticket >= 121.5:
        return " TIENDA PREMIUM" if upf >= 1.9 else "TIENDA CON TICKET ALTO"
    return "TIENDA DEBE MEJORAR TICKET" if upf >= 1.9 else "OPORTUNIDAD DE MEJORAR"


# Empresa cuyas bodegas son tiendas. ENV01 solo tiene 6 bodegas, todas de
# logistica y administracion ("CENTRO LOG UIO", "CENTR LOG DURAN"), y comparten
# `Codigo_Local` con tiendas reales: sin este filtro la sucursal 004 suma 1116
# lineas en vez de 166.
CORP_TIENDAS = "NVC01"

# El ERP devuelve 200 filas cuando no se le pasa limite, y no avisa. El maestro
# de bodegas tiene 329.
LIMITE_ERP = 5000

_RE_PREFIJO_3 = re.compile(r"^\s*(\d{3})")

# --- Credito directo (CrediNovi / Banco Solidario) -------------------------
# El tipo de pago lo teclea el cajero y sale escrito de trece formas distintas
# ("CREDINOVI", "CREDINOV1", "CREDNOVI", "CrediNovi", "CREDINOVI 1",
# "CREDITO DIRECTO", "BSOL", "BANCO SOLIDARIO"...). Ademas hay dos campos
# intercambiables: unos lo escriben en BANK_O_CC_TYPE y otros en
# NOMBRE_BANCO_O_TIPO_TC. Y algunos no escriben nada: esos solo se reconocen
# por el numero de operacion de CrediNovi, que empieza en 107.
_RE_OPERACION_CREDINOVI = re.compile(r"107\d{3}")
FORMA_PAGO_CREDITO = "Otros"


def _solo_letras(texto) -> str:
    return re.sub(r"[^A-Z]", "", str(texto or "").upper())


def es_credito_directo(tipo_banco, nombre_banco) -> bool:
    """Decide si un pago 'Otros' es credito directo.

    Valida contra la hoja D.CREDINOVI de agosto: 42 de 42 cobros, sin falsos
    positivos entre los 328 pagos 'Otros' del periodo.
    """
    texto = _solo_letras(tipo_banco) + " " + _solo_letras(nombre_banco)
    if "CRED" in texto and ("NOV" in texto or "DIRECTO" in texto):
        return True
    if "SOLIDARIO" in texto or "BSOL" in texto:
        return True
    # Sin etiqueta: el numero de operacion es el unico rastro.
    return bool(_RE_OPERACION_CREDINOVI.fullmatch(str(nombre_banco or "").strip()))


def derivar_sucursal(ware_name: str, codigo_local: str) -> Optional[str]:
    """Deduce a que tienda pertenece una bodega.

    El nombre de la bodega de tienda ya trae el numero delante ("008 CITY MALL"),
    que es exactamente la etiqueta que usa el reporte manual. Si no lo trae, se
    cae al `Codigo_Local` cuando es un codigo de sucursal de 3 digitos.

    Devuelve None cuando no se puede deducir: esas bodegas quedan fuera del
    reporte hasta que alguien les asigne una tienda a mano.
    """
    m = _RE_PREFIJO_3.match(ware_name or "")
    if m:
        return m.group(1)
    local = (codigo_local or "").strip()
    return local if re.fullmatch(r"\d{3}", local) else None


def _rango_periodo(periodo: str) -> tuple:
    anio, mes = (int(p) for p in periodo.split("-"))
    dias = calendar.monthrange(anio, mes)[1]
    return (datetime.date(anio, mes, 1), datetime.date(anio, mes, dias), dias)


def _cumplimiento(real: float, meta: float, peso: float) -> float:
    """Ratio topado al peso.

    Replica el IFS del Excel: por debajo de la meta paga proporcional, en la
    meta o por encima paga el peso completo. Sin meta no puntua (en el Excel
    esa celda daba #DIV/0!).
    """
    if meta is None or meta <= 0:
        return 0.0
    return min(real / meta, 1.0) * peso


class KpiService:
    """Reporte de Seguimiento KPI por sucursal.

    Sustituye el armado manual del Excel: las categorias salen de cruzar la
    vista de ventas contra `kpi_producto_cat`, y las metas de `kpi_meta`.
    """

    def obtener_seguimiento(self, periodo: str, corte: Optional[str] = None,
                            db: Optional[Session] = None) -> dict:
        inicio, fin_mes, dias_mes = _rango_periodo(periodo)

        cerrar = False
        if db is None:
            db = SessionLocal()
            cerrar = True
        try:
            with db.get_bind().connect() as conn:
                if corte:
                    fin = datetime.date.fromisoformat(corte)
                else:
                    # El corte real es el ultimo dia sincronizado, no "hoy": el
                    # sync es manual 4 veces al dia y el mes casi nunca esta
                    # completo cuando se corre el reporte.
                    fin = conn.execute(
                        text("SELECT MAX(fecha) FROM view_ventas_espejo_reporte "
                             "WHERE fecha BETWEEN :i AND :f"),
                        {"i": inicio, "f": fin_mes}).scalar() or inicio
                fin = min(fin, fin_mes)
                params = {"i": inicio, "f": fin}

                # El codigo de la vista conserva el sufijo de empresa
                # ("1CENV153-NVC01") pero el catalogo usa el codigo pelado, asi
                # que se quita antes de cruzar. Sin esto el join no pega nunca.
                # La sucursal sale del mapeo de bodegas, no de `v.sucursal`
                # (codigo_local): ese codigo lo comparten la tienda y las
                # bodegas de logistica de ENV, que no son venta de tienda.
                # El JOIN deja fuera las bodegas sin mapear (mayoristas,
                # e-commerce, ROBO/PERDIDAS), igual que el reporte manual.
                filas = conn.execute(text("""
                    SELECT COALESCE(b.sucursal_override, b.sucursal) AS sucursal,
                           UPPER(TRIM(c.cat)) AS cat,
                           COALESCE(SUM(v.cantidad), 0) AS unidades,
                           COALESCE(SUM(v.total_linea), 0) AS monto
                    FROM view_ventas_espejo_reporte v
                    JOIN kpi_bodega b ON b.ware_code = v.bodega_codigo
                    JOIN kpi_producto_cat c
                      ON UPPER(TRIM(c.codigo)) =
                         UPPER(regexp_replace(v.codigo, '-(NVC01|ENV01)$', ''))
                    WHERE v.fecha BETWEEN :i AND :f
                      AND COALESCE(b.sucursal_override, b.sucursal) IS NOT NULL
                    GROUP BY 1, 2
                """), params).mappings().all()

                metas = conn.execute(text(
                    "SELECT sucursal, kpi, meta FROM kpi_meta WHERE periodo = :p"),
                    {"p": periodo}).mappings().all()
                manuales = conn.execute(text(
                    "SELECT sucursal, kpi, valor FROM kpi_valor_manual "
                    "WHERE periodo = :p"), {"p": periodo}).mappings().all()
                # El cobro ya trae su sucursal (CODIGO_TIENDA), asi que este KPI
                # no pasa por el mapeo de bodegas.
                cobros = conn.execute(text("""
                    SELECT sucursal, COALESCE(SUM(valor), 0) AS monto
                    FROM kpi_cobro_credito
                    WHERE fecha BETWEEN :i AND :f AND sucursal IS NOT NULL
                    GROUP BY 1
                """), params).mappings().all()
                # Rentabilidad: margen sobre TODA la venta de la tienda, sin
                # filtrar por categoria (a diferencia del resto de KPIs).
                margenes = conn.execute(text("""
                    SELECT COALESCE(b.sucursal_override, b.sucursal) AS sucursal,
                           COALESCE(SUM(v.total_linea), 0) AS venta,
                           COALESCE(SUM(v.costo_total), 0) AS costo
                    FROM view_ventas_espejo_reporte v
                    JOIN kpi_bodega b ON b.ware_code = v.bodega_codigo
                    WHERE v.fecha BETWEEN :i AND :f
                      AND COALESCE(b.sucursal_override, b.sucursal) IS NOT NULL
                    GROUP BY 1
                """), params).mappings().all()
                sucursales = conn.execute(text(
                    "SELECT codigo, nombre, supervisor, marca, ciudad "
                    "FROM kpi_sucursal WHERE activa = 'SI'")).mappings().all()

            reales = {}
            cat_a_kpi = {v.upper().strip(): k for k, v in CATS_VENTAS.items()}
            for f in filas:
                kpi = cat_a_kpi.get(f["cat"])
                if not kpi:
                    continue     # categoria fuera de los KPIs (ACCESORIO, etc.)
                medida = KPIS[kpi]["medida"]
                valor = float(f["monto"] if medida == "monto" else f["unidades"])
                reales[(f["sucursal"], kpi)] = valor
            for m in manuales:
                reales[(m["sucursal"], m["kpi"])] = float(m["valor"] or 0)
            for c in cobros:
                reales[(c["sucursal"], "credito_directo")] = float(c["monto"] or 0)
            for r in margenes:
                venta = float(r["venta"] or 0)
                if venta:
                    reales[(r["sucursal"], "rentabilidad")] = (
                        venta - float(r["costo"] or 0)) / venta

            metas_map = {(m["sucursal"], m["kpi"]): float(m["meta"] or 0)
                         for m in metas}

            dias_corte = (fin - inicio).days + 1
            resultado = []
            for s in sucursales:
                cod = s["codigo"]
                detalle, total = [], 0.0
                for kpi, cfg in KPIS.items():
                    real = reales.get((cod, kpi), 0.0)
                    meta = metas_map.get((cod, kpi))
                    aporte = _cumplimiento(real, meta or 0, cfg["peso"])
                    total += aporte
                    detalle.append({
                        "kpi": kpi,
                        "label": cfg["label"],
                        "origen": cfg["origen"],
                        "medida": cfg["medida"],
                        "peso": cfg["peso"],
                        "meta": meta,
                        "real": real,
                        "cumplimiento": (real / meta) if meta else None,
                        "aporte": round(aporte, 6),
                        "proyeccion": round(real / dias_corte * dias_mes, 2)
                        if dias_corte else None,
                    })
                resultado.append({
                    "sucursal": cod,
                    "nombre": s["nombre"],
                    "supervisor": s["supervisor"],
                    "marca": s["marca"],
                    "ciudad": s["ciudad"],
                    "total_kpi": round(total, 6),
                    "sin_metas": all(d["meta"] is None for d in detalle),
                    "detalle": detalle,
                })

            resultado.sort(key=lambda r: r["total_kpi"], reverse=True)
            return {
                "periodo": periodo,
                "inicio": inicio.isoformat(),
                "corte": fin.isoformat(),
                "dias_corte": dias_corte,
                "dias_mes": dias_mes,
                "peso_total": round(sum(k["peso"] for k in KPIS.values()), 4),
                "sucursales": resultado,
            }
        except Exception as e:
            logging.error(f"Error al construir el seguimiento KPI: {e}")
            raise
        finally:
            if cerrar:
                db.close()

    def sincronizar_bodegas(self, repository, env: Optional[str] = None,
                            db: Optional[Session] = None) -> dict:
        """Trae el maestro de bodegas del ERP y recalcula el mapeo a sucursal.

        No pisa `sucursal_override`: las correcciones hechas a mano sobreviven a
        cada sincronizacion.
        """
        filas = repository.ejecutar_consulta(
            repository.obtener_token(env=env),
            "WARE_CODE,WARE_NAME,Codigo_Local,INACTIVE,CORP",
            "INVT_Bodegas_Lista", limit=LIMITE_ERP, env=env)
        if not filas:
            raise ValueError("El ERP no devolvio bodegas (revisar permisos del servicio).")

        cerrar = False
        if db is None:
            db = SessionLocal()
            cerrar = True
        try:
            mapeadas = 0
            for f in filas:
                code = str(f.get("WARE_CODE") or "").strip()
                if not code:
                    continue
                nombre = str(f.get("WARE_NAME") or "").strip()
                local = str(f.get("Codigo_Local") or "").strip()
                corp = str(f.get("CORP") or "").strip()
                sucursal = derivar_sucursal(nombre, local) if corp == CORP_TIENDAS else None
                mapeadas += 1 if sucursal else 0
                db.execute(text("""
                    INSERT INTO kpi_bodega
                        (ware_code, ware_name, codigo_local, corp, inactiva, sucursal)
                    VALUES (:c, :n, :l, :corp, :inact, :suc)
                    ON CONFLICT (ware_code) DO UPDATE SET
                        ware_name = EXCLUDED.ware_name,
                        codigo_local = EXCLUDED.codigo_local,
                        corp = EXCLUDED.corp,
                        inactiva = EXCLUDED.inactiva,
                        sucursal = EXCLUDED.sucursal,
                        updated_at = NOW()
                """), {"c": code, "n": nombre, "l": local, "corp": corp,
                       "inact": bool(f.get("INACTIVE")), "suc": sucursal})
            # El kardex a veces guarda un `Codigo_Local` en el campo de bodega
            # ("POT", "QUI", "PRI"), y esos codigos NO existen en el maestro. Se
            # registran igual, sin sucursal, para que aparezcan en la lista de
            # pendientes en vez de desaparecer del reporte en silencio.
            huerfanas = db.execute(text("""
                INSERT INTO kpi_bodega (ware_code, ware_name)
                SELECT DISTINCT v.bodega_codigo, '(no esta en el maestro)'
                FROM view_ventas_espejo_reporte v
                WHERE v.bodega_codigo IS NOT NULL AND v.bodega_codigo <> ''
                  AND NOT EXISTS (
                      SELECT 1 FROM kpi_bodega b WHERE b.ware_code = v.bodega_codigo)
                ON CONFLICT (ware_code) DO NOTHING
            """)).rowcount or 0
            db.commit()
            return {"bodegas": len(filas), "mapeadas": mapeadas,
                    "sin_mapear": len(filas) - mapeadas,
                    "fuera_del_maestro": huerfanas}
        finally:
            if cerrar:
                db.close()

    def sincronizar_cobros(self, repository, inicio: str, fin: str,
                           env: Optional[str] = None,
                           db: Optional[Session] = None) -> dict:
        """Trae los cobros de credito directo del ERP para un rango de fechas.

        El filtro de fecha NO es opcional: estas tablas devuelven 3000 filas como
        maximo ignorando el `limit`, y sin `where` entregan las mas antiguas
        (registros de 2018) sin ningun aviso.
        """
        filas = repository.ejecutar_consulta(
            repository.obtener_token(env=env),
            "CODIGO_COBRO,BANK_O_CC_TYPE,NOMBRE_BANCO_O_TIPO_TC,CODIGO_TIENDA,"
            "VALOR_DE_PAGO,FECHA_PAGO",
            "CLNT_Cobro_FormaDePago",
            where=(f"FORMA_DE_PAGO='{FORMA_PAGO_CREDITO}' "
                   f"AND FECHA_PAGO>='{inicio}' AND FECHA_PAGO<='{fin}'"),
            limit=LIMITE_ERP, env=env)

        cerrar = False
        if db is None:
            db = SessionLocal()
            cerrar = True
        try:
            guardados = 0
            for f in filas:
                tipo, nombre = f.get("BANK_O_CC_TYPE"), f.get("NOMBRE_BANCO_O_TIPO_TC")
                if not es_credito_directo(tipo, nombre):
                    continue
                codigo = str(f.get("CODIGO_COBRO") or "").strip()
                if not codigo:
                    continue
                db.execute(text("""
                    INSERT INTO kpi_cobro_credito
                        (codigo_cobro, sucursal, fecha, valor, tipo_crudo)
                    VALUES (:c, :s, :f, :v, :t)
                    ON CONFLICT (codigo_cobro) DO UPDATE SET
                        sucursal = EXCLUDED.sucursal, fecha = EXCLUDED.fecha,
                        valor = EXCLUDED.valor, tipo_crudo = EXCLUDED.tipo_crudo,
                        updated_at = NOW()
                """), {"c": codigo,
                       "s": str(f.get("CODIGO_TIENDA") or "").strip() or None,
                       "f": str(f.get("FECHA_PAGO"))[:10],
                       "v": float(f.get("VALOR_DE_PAGO") or 0),
                       "t": f"{tipo or ''}|{nombre or ''}"[:160]})
                guardados += 1
            db.commit()
            return {"inicio": inicio, "fin": fin, "pagos_otros": len(filas),
                    "credito_directo": guardados}
        finally:
            if cerrar:
                db.close()

    def importar_excel(self, contenido: bytes, periodo: str,
                       db: Optional[Session] = None,
                       nombre_archivo: Optional[str] = None) -> dict:
        """Carga sucursales, catalogo y metas desde el Excel armado a mano.

        Es la unica via para sembrar sin entrar por SSH al servidor. Idempotente:
        volver a subir el mismo archivo actualiza en vez de duplicar.

        Lee tres hojas:
          PRESUPUESTO -> sucursal, marca, ciudad, supervisor y meta de tienda
          POND!J:L    -> catalogo producto -> categoria (el VLOOKUP del Excel)
          RESUMEN KPI -> meta de cada KPI por sucursal
        """
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
        faltan = [h for h in ("PRESUPUESTO", "POND", "RESUMEN KPI")
                  if h not in wb.sheetnames]
        if faltan:
            raise ValueError(f"Al archivo le faltan hojas: {', '.join(faltan)}")

        sucursales, metas = {}, []
        for fila in wb["PRESUPUESTO"].iter_rows(min_row=2, values_only=True):
            m = RE_SUCURSAL_EXCEL.match(str(fila[1] or ""))
            if not m:
                continue
            sucursales[m.group(1)] = {
                "codigo": m.group(1), "nombre": m.group(2),
                "marca": fila[2], "ciudad": fila[3], "supervisor": fila[4],
            }
            if isinstance(fila[5], (int, float)):      # META DE TIENDA del mes
                metas.append((m.group(1), KPI_VENTA_TIENDA, float(fila[5])))

        catalogo = {}
        for fila in wb["POND"].iter_rows(min_row=2, min_col=10, max_col=12,
                                         values_only=True):
            codigo, producto, cat = fila
            if codigo and cat:
                catalogo[str(codigo).strip().upper()] = (
                    str(producto or "").strip(), str(cat).strip().upper())

        for fila in wb["RESUMEN KPI"].iter_rows(min_row=3, values_only=True):
            m = RE_SUCURSAL_EXCEL.match(str(fila[1] or ""))
            if not m:
                continue
            for col, kpi in COL_META_EXCEL.items():
                if col < len(fila) and isinstance(fila[col], (int, float)):
                    metas.append((m.group(1), kpi, float(fila[col])))

        if not sucursales:
            raise ValueError("No se reconocio ninguna sucursal en la hoja PRESUPUESTO.")

        # La hoja BASE trae Bodega y SUCURSAL en la misma fila: es el unico lugar
        # donde consta a que tienda pertenecen las bodegas ADMIN ("ADMIN NV
        # BOMBOL" -> 164), que no se puede deducir de ningun campo del ERP.
        overrides = {}
        if "BASE" in wb.sheetnames:
            for fila in wb["BASE"].iter_rows(min_row=2, values_only=True):
                if not fila or not fila[0]:
                    continue
                m = RE_SUCURSAL_EXCEL.match(str(fila[6] or ""))
                bodega = str(fila[5] or "").strip().upper()
                if m and bodega:
                    overrides[bodega] = m.group(1)

        cerrar = False
        if db is None:
            db = SessionLocal()
            cerrar = True
        try:
            for s in sucursales.values():
                db.execute(text("""
                    INSERT INTO kpi_sucursal (codigo, nombre, supervisor, marca, ciudad, activa)
                    VALUES (:codigo, :nombre, :supervisor, :marca, :ciudad, 'SI')
                    ON CONFLICT (codigo) DO UPDATE SET
                        nombre = EXCLUDED.nombre, supervisor = EXCLUDED.supervisor,
                        marca = EXCLUDED.marca, ciudad = EXCLUDED.ciudad
                """), s)
            for codigo, (producto, cat) in catalogo.items():
                db.execute(text("""
                    INSERT INTO kpi_producto_cat (codigo, cat, producto)
                    VALUES (:codigo, :cat, :producto)
                    ON CONFLICT (codigo) DO UPDATE SET
                        cat = EXCLUDED.cat, producto = EXCLUDED.producto
                """), {"codigo": codigo, "cat": cat, "producto": producto})
            for sucursal, kpi, meta in metas:
                db.execute(text("""
                    INSERT INTO kpi_meta (periodo, sucursal, kpi, meta)
                    VALUES (:p, :s, :k, :m)
                    ON CONFLICT (periodo, sucursal, kpi)
                    DO UPDATE SET meta = EXCLUDED.meta, updated_at = NOW()
                """), {"p": periodo, "s": sucursal, "k": kpi, "m": meta})
            # Solo se marcan las bodegas que el ERP ya conoce: si el codigo del
            # Excel no existe en el maestro, es otro identificador y no sirve.
            # El mismo archivo sirve de plantilla para generar el reporte con su
            # formato exacto (colores, fuentes, anchos), en vez de replicarlo.
            db.execute(text("""
                INSERT INTO kpi_plantilla (id, nombre, archivo)
                VALUES (1, :n, :a)
                ON CONFLICT (id) DO UPDATE SET
                    nombre = EXCLUDED.nombre, archivo = EXCLUDED.archivo,
                    updated_at = NOW()
            """), {"n": nombre_archivo, "a": contenido})

            aplicados = 0
            for bodega, sucursal in overrides.items():
                # El Excel escribe la bodega sin ceros a la izquierda ("16" donde
                # el ERP dice "016"), asi que se compara de las dos formas.
                # Nunca se le asigna sucursal a una bodega que no es de tienda:
                # el Excel escribe "3" y al rellenar a "003" coincide con la
                # bodega de logistica de ENV, que reventaba la sucursal 003.
                r = db.execute(text("""
                    UPDATE kpi_bodega SET sucursal_override = :s, updated_at = NOW()
                    WHERE (UPPER(ware_code) = :b OR UPPER(ware_code) = LPAD(:b, 3, '0'))
                      AND COALESCE(corp, :corp) = :corp
                      AND COALESCE(sucursal, '') <> :s
                """), {"b": bodega, "s": sucursal, "corp": CORP_TIENDAS})
                aplicados += r.rowcount or 0
            db.commit()
        finally:
            if cerrar:
                db.close()

        return {"periodo": periodo, "sucursales": len(sucursales),
                "productos": len(catalogo), "metas": len(metas),
                "bodegas_corregidas": aplicados}

    def obtener_lineas(self, inicio: str, fin: str, solo_categorizadas: bool = True,
                       db: Optional[Session] = None) -> list:
        """Lineas de venta con su categoria de KPI, para las hojas de detalle.

        Con `solo_categorizadas=False` devuelve tambien lo que no puntua en
        ningun KPI (accesorios, marcas de terceros), que es lo que la hoja BASE
        del Excel muestra con #N/A.
        """
        cerrar = False
        if db is None:
            db = SessionLocal()
            cerrar = True
        try:
            join = "JOIN" if solo_categorizadas else "LEFT JOIN"
            with db.get_bind().connect() as conn:
                filas = conn.execute(text(f"""
                    SELECT v.factura_final, v.fecha, v.codigo, v.producto,
                           v.unidad, v.grupo, v.subgrupo, v.cantidad,
                           v.precio_venta, v.total_linea,
                           v.bodega_codigo,
                           v.codigo_cliente AS codigo_vendedor,
                           v.nombre_cliente,
                           -- El reporte manual trae esta columna fija en TIENDA:
                           -- las bodegas que no son de tienda ya quedaron fuera
                           -- por el mapeo, asi que aqui todo lo que pasa es tienda.
                           'TIENDA' AS canal,
                           COALESCE(b.sucursal_override, b.sucursal) AS sucursal,
                           TRIM(CONCAT(COALESCE(b.sucursal_override, b.sucursal),
                                       ' ', COALESCE(s.nombre, ''))) AS sucursal_larga,
                           s.nombre AS sucursal_nombre, s.supervisor,
                           UPPER(TRIM(c.cat)) AS cat
                    FROM view_ventas_espejo_reporte v
                    JOIN kpi_bodega b ON b.ware_code = v.bodega_codigo
                    {join} kpi_producto_cat c
                      ON UPPER(TRIM(c.codigo)) =
                         UPPER(regexp_replace(v.codigo, '-(NVC01|ENV01)$', ''))
                    LEFT JOIN kpi_sucursal s
                      ON s.codigo = COALESCE(b.sucursal_override, b.sucursal)
                    WHERE v.fecha BETWEEN :i AND :f
                      AND COALESCE(b.sucursal_override, b.sucursal) IS NOT NULL
                    ORDER BY v.fecha, v.factura_final
                """), {"i": inicio, "f": fin}).mappings().all()
            return [dict(f) for f in filas]
        finally:
            if cerrar:
                db.close()

    def obtener_presupuesto(self, periodo: str, corte: Optional[str] = None,
                            db: Optional[Session] = None) -> list:
        """Venta total por tienda contra su meta mensual: la hoja PRESUPUESTO.

        La proyeccion extrapola lo vendido hasta el corte al mes completo, igual
        que la formula del Excel.
        """
        inicio, fin_mes, dias_mes = _rango_periodo(periodo)
        cerrar = False
        if db is None:
            db = SessionLocal()
            cerrar = True
        try:
            with db.get_bind().connect() as conn:
                if corte:
                    fin = datetime.date.fromisoformat(corte)
                else:
                    fin = conn.execute(
                        text("SELECT MAX(fecha) FROM view_ventas_espejo_reporte "
                             "WHERE fecha BETWEEN :i AND :f"),
                        {"i": inicio, "f": fin_mes}).scalar() or inicio
                fin = min(fin, fin_mes)

                filas = conn.execute(text("""
                    SELECT COALESCE(b.sucursal_override, b.sucursal) AS sucursal,
                           COALESCE(SUM(v.total_linea), 0) AS venta,
                           COUNT(DISTINCT v.factura_final) AS facturas,
                           COALESCE(SUM(v.cantidad), 0) AS unidades
                    FROM view_ventas_espejo_reporte v
                    JOIN kpi_bodega b ON b.ware_code = v.bodega_codigo
                    WHERE v.fecha BETWEEN :i AND :f
                      AND COALESCE(b.sucursal_override, b.sucursal) IS NOT NULL
                    GROUP BY 1
                """), {"i": inicio, "f": fin}).mappings().all()
                metas = conn.execute(text(
                    "SELECT sucursal, meta FROM kpi_meta "
                    "WHERE periodo = :p AND kpi = :k"),
                    {"p": periodo, "k": KPI_VENTA_TIENDA}).mappings().all()
                sucursales = conn.execute(text(
                    "SELECT codigo, nombre, supervisor, marca, ciudad "
                    "FROM kpi_sucursal WHERE activa = 'SI'")).mappings().all()

            ventas = {f["sucursal"]: f for f in filas}
            metas_map = {m["sucursal"]: float(m["meta"] or 0) for m in metas}
            dias_corte = (fin - inicio).days + 1

            out = []
            for s in sucursales:
                v = ventas.get(s["codigo"])
                venta = float(v["venta"]) if v else 0.0
                facturas = int(v["facturas"]) if v else 0
                unidades = float(v["unidades"]) if v else 0.0
                meta = metas_map.get(s["codigo"])
                promedio_dia = venta / dias_corte if dias_corte else 0.0
                out.append({
                    "sucursal": s["codigo"], "nombre": s["nombre"],
                    "supervisor": s["supervisor"], "marca": s["marca"],
                    "ciudad": s["ciudad"],
                    "meta": meta, "venta": round(venta, 2),
                    "facturas": facturas, "unidades": unidades,
                    "ticket_promedio": round(venta / facturas, 2) if facturas else 0.0,
                    "unidades_por_factura": round(unidades / facturas, 2) if facturas else 0.0,
                    "venta_promedio_dia": round(promedio_dia, 2),
                    "proyeccion": round(promedio_dia * dias_mes, 2),
                    "cumplimiento": round(venta / meta, 4) if meta else None,
                })
            out.sort(key=lambda r: r["venta"], reverse=True)
            return out
        finally:
            if cerrar:
                db.close()
