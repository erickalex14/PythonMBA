import requests
import logging
import pandas as pd
import json
import datetime
import time
import re
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configuracion del registro de eventos
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Constantes de configuracion
BASE_URL = "http://181.198.104.181:8020"
URL_LOGIN = f"{BASE_URL}/ws2_mba3_serv_/login_servicio"
URL_CONSULTA = f"{BASE_URL}/ws2_mba3_serv_Consultas_Externas_/"

CODIGO_SERVICIO = "SERIALES"
PASSWORD_SERVICIO = "***REDACTED***"


def solicitar_fecha(mensaje_prompt):
    while True:
        fecha_str = input(mensaje_prompt).strip()
        try:
            fecha_obj = datetime.datetime.strptime(fecha_str, "%Y-%m-%d")
            return fecha_obj
        except ValueError:
            print("Error: Formato de fecha incorrecto. Utilice YYYY-MM-DD (ejemplo: 2026-06-01).")


def solicitar_rango_fechas():
    print("\n--- Generador Masivo de Reporte de Liquidaciones Consolidado MBA3 ---")
    print("Por favor, ingrese el rango de fechas para la consulta.\n")

    while True:
        fecha_inicio = solicitar_fecha("Ingrese la fecha de INICIO (YYYY-MM-DD): ")
        fecha_fin = solicitar_fecha("Ingrese la fecha de FIN (YYYY-MM-DD): ")

        if fecha_inicio <= fecha_fin:
            print(
                f"\nRango seleccionado: Desde {fecha_inicio.strftime('%Y-%m-%d')} hasta {fecha_fin.strftime('%Y-%m-%d')}\n")
            return fecha_inicio, fecha_fin
        else:
            print("\nError: La fecha de inicio no puede ser posterior a la fecha de fin. Intente nuevamente.\n")


def obtener_token():
    logging.info("Solicitando nuevas credenciales de sesion al ERP...")
    headers = {"Content-Type": "application/json"}
    payload = {"codigo": CODIGO_SERVICIO, "pwd": PASSWORD_SERVICIO}

    try:
        response = requests.post(URL_LOGIN, json=payload, headers=headers, timeout=15)
        response.raise_for_status()
        datos = response.json()
        token = datos.get("jwt")
        if token:
            logging.info("Autenticacion exitosa. Nuevo token asignado.")
            return token
        else:
            logging.error("Respuesta de autenticacion sin token JWT. Detalle: %s", datos)
            return None
    except requests.exceptions.RequestException as e:
        logging.error("Fallo de red al intentar autenticar: %s", e)
        return None


def ejecutar_consulta_diaria(token, fecha_consulta):
    headers = {"Authorization": token}
    fecha_str = fecha_consulta.strftime('%Y-%m-%d')

    # Condicion estricta con LIQUIDACION_ESTADO = True
    condicion_where = f"CORP = 'NVC01' AND LIQUIDACION_ESTADO = True AND LIQUIDACION_FECHA = '{fecha_str}'"
    cols_principal = "CORP,LIQUIDACION_FECHA,OBSERVACIONES,ANTES_TOTAL_1,ANTES_TOTAL_2,ANTES_TOTAL_3,DESPUES_TOTAL_1,DESPUES_TOTAL_2,DESPUES_TOTAL_3,LIQUIDACION_ESTADO,LIQUIDACION_ID_CORP"

    payload = {
        "select": cols_principal,
        "from": "PROV_Liquidaciones_Principal",
        "where": condicion_where,
        "limit": "50000"
    }

    try:
        response = requests.post(URL_CONSULTA, headers=headers, data=payload, timeout=150)

        if response.status_code == 401:
            return "UNAUTHORIZED"

        response.raise_for_status()
        datos = response.json()

        if isinstance(datos, dict):
            if datos.get("codigo") == "009":
                return []
            elif "error" in datos or "codigo" in datos:
                logging.error("El servidor rechazo la consulta. Detalle: %s", datos)
                return []

        if isinstance(datos, dict):
            return [datos]
        elif isinstance(datos, list):
            return datos
        else:
            return []

    except requests.exceptions.Timeout:
        return "TIMEOUT"
    except requests.exceptions.RequestException as e:
        logging.error("Error de conexion general: %s", e)
        return "ERROR"
    except json.JSONDecodeError:
        logging.error("Fallo en la decodificacion JSON.")
        return "ERROR"


def ejecutar_consulta_tabla(token, columnas, tabla, condicion_where=None, limite=None):
    headers = {"Authorization": token}
    payload = {"select": columnas, "from": tabla}

    if condicion_where:
        payload["where"] = condicion_where
    if limite:
        payload["limit"] = str(limite)

    try:
        response = requests.post(URL_CONSULTA, headers=headers, data=payload, timeout=150)
        if response.status_code == 401:
            return "UNAUTHORIZED"

        response.raise_for_status()
        datos = response.json()

        if isinstance(datos, dict):
            if datos.get("codigo") == "009":
                return []
            elif "error" in datos or "codigo" in datos:
                return "ERROR"
        return datos
    except requests.exceptions.Timeout:
        return "TIMEOUT"
    except Exception:
        return "ERROR"


def procesar_cruce_datos(token_inicial, fecha_inicio, fecha_fin):
    # =========================================================
    # 1. EXTRACCION DE CABECERAS EN LOTES DIARIOS
    # =========================================================
    registros_cabecera = []
    fecha_actual = fecha_inicio
    dias_totales = (fecha_fin - fecha_inicio).days + 1
    dia_contador = 1
    token_actual = token_inicial

    logging.info("Fase 1/3: Descargando datos de cabecera de liquidaciones. Dias a procesar: %d", dias_totales)

    while fecha_actual <= fecha_fin:
        fecha_str = fecha_actual.strftime('%Y-%m-%d')

        max_intentos = 3
        intento_actual = 1
        exito = False

        while intento_actual <= max_intentos and not exito:
            resultado = ejecutar_consulta_diaria(token_actual, fecha_actual)

            if resultado == "UNAUTHORIZED":
                logging.warning("La sesion expiro. Renovando token automaticamente...")
                nuevo_token = obtener_token()
                if nuevo_token:
                    token_actual = nuevo_token
                    time.sleep(1)
                    continue
                else:
                    logging.critical("Fallo critico al renovar credenciales. Abortando extraccion.")
                    return None, token_actual

            elif resultado == "TIMEOUT":
                intento_actual += 1
                time.sleep(5)
            elif resultado == "ERROR":
                break
            else:
                if resultado:
                    registros_cabecera.extend(resultado)
                exito = True

        fecha_actual += datetime.timedelta(days=1)
        dia_contador += 1

    if not registros_cabecera:
        logging.warning("Sin resultados en tabla principal para este rango de fechas. Proceso abortado.")
        return None, token_actual

    df_principal = pd.DataFrame(registros_cabecera)
    logging.info("Total registros obtenidos en cabecera: %d", len(df_principal))

    # Normalizacion de IDs relacionales
    df_principal['LIQUIDACION_ID_CORP'] = df_principal['LIQUIDACION_ID_CORP'].astype(str).str.replace(r'\.0$', '',
                                                                                                      regex=True).str.strip().str.upper()

    # =========================================================
    # 2. EXTRACCION MASIVA DE PRODUCTOS (Volvemos a RAM con alto limite)
    # =========================================================
    cols_productos = "FACTURA_ID_CORP,IdRecepcionRelacionada,VALOR_TOTAL_CIF,VALOR_SUBTOTAL_CIF,VALOR_ANTES_1,VALOR_ANTES_2,VALOR_ANTES_3,VALOR_DESPUES_1,VALOR_DESPUES_2,VALOR_DESPUES_3,PARTIDA_ID_CORP,PRODUCTO_ID_CORP,LIQUIDACION_ID,CANTIDAD,PRECIO,TOTAL,VALOR_TOTAL_CIF_MANUAL,VALOR_TOTAL_CIF_UNIDAD,LIQUIDACION_ID_CORP"

    logging.info("Fase 2/3: Descargando tabla de productos de forma MASIVA (Limite robusto: 100,000)...")

    while True:
        # Quitamos la clausula IN que rompia el gateway y pedimos todo directo en un solo bloque estable
        datos_productos = ejecutar_consulta_tabla(token_actual, cols_productos, "PROV_Liquidaciones_Productos",
                                                  limite=100000)

        if datos_productos == "UNAUTHORIZED":
            logging.warning("Token expirado al descargar productos. Renovando token...")
            nuevo_token = obtener_token()
            if nuevo_token:
                token_actual = nuevo_token
                time.sleep(1)
                continue
            else:
                return None, token_actual
        elif datos_productos == "TIMEOUT" or datos_productos == "ERROR" or not datos_productos:
            logging.warning("No se pudo obtener la tabla secundaria o vino vacia.")
            return None, token_actual
        else:
            break

    df_productos = pd.DataFrame(datos_productos if isinstance(datos_productos, list) else [datos_productos])
    df_productos['LIQUIDACION_ID_CORP'] = df_productos['LIQUIDACION_ID_CORP'].astype(str).str.replace(r'\.0$', '',
                                                                                                      regex=True).str.strip().str.upper()
    logging.info("Total registros obtenidos en productos: %d", len(df_productos))

    # =========================================================
    # 3. OPERACION INNER JOIN EN MEMORIA RAM
    # =========================================================
    logging.info("Fase 3/3: Ejecutando cruce relacional (INNER JOIN) en memoria RAM.")
    df_cruce = pd.merge(df_principal, df_productos, on='LIQUIDACION_ID_CORP', how='inner')

    columnas_finales = [
        "CORP", "LIQUIDACION_FECHA", "OBSERVACIONES", "FACTURA_ID_CORP",
        "IdRecepcionRelacionada", "ANTES_TOTAL_1", "ANTES_TOTAL_2", "ANTES_TOTAL_3",
        "DESPUES_TOTAL_1", "DESPUES_TOTAL_2", "DESPUES_TOTAL_3", "VALOR_TOTAL_CIF",
        "VALOR_SUBTOTAL_CIF", "VALOR_ANTES_1", "VALOR_ANTES_2", "VALOR_ANTES_3",
        "VALOR_DESPUES_1", "VALOR_DESPUES_2", "VALOR_DESPUES_3", "PARTIDA_ID_CORP",
        "PRODUCTO_ID_CORP", "LIQUIDACION_ID", "CANTIDAD", "PRECIO", "TOTAL",
        "VALOR_TOTAL_CIF_MANUAL", "VALOR_TOTAL_CIF_UNIDAD", "LIQUIDACION_ESTADO",
        "LIQUIDACION_ID_CORP"
    ]

    columnas_disponibles = [col for col in columnas_finales if col in df_cruce.columns]
    df_final = df_cruce[columnas_disponibles]

    logging.info("Operacion de cruce finalizada. Lineas listas para exportacion: %d", len(df_final))
    return df_final, token_actual


def exportar_a_excel_empresarial(df_datos, fecha_inicio, fecha_fin):
    logging.info("Iniciando modulo de renderizado de reporte Excel corporativo con saneamiento de datos.")

    if df_datos is None or df_datos.empty:
        logging.warning("Conjunto de datos vacio. Se omite la generacion del archivo.")
        return

    try:
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

        def sanitizar_excel(valor):
            if pd.isna(valor):
                return valor
            val_str = str(valor).strip()
            val_str = ILLEGAL_CHARACTERS_RE.sub("", val_str)
            if len(val_str) > 32700:
                return val_str[:32700] + "... [TRUNCADO POR LIMITE DE EXCEL]"
            return val_str

        for col in df_datos.select_dtypes(include=['object']).columns:
            df_datos[col] = df_datos[col].apply(sanitizar_excel)

        # Nombre estricto: LIQUIDACIONES (YYYY-MM-DD al YYYY-MM-DD).xlsx
        str_inicio = fecha_inicio.strftime('%Y-%m-%d')
        str_fin = fecha_fin.strftime('%Y-%m-%d')
        archivo_final = f"LIQUIDACIONES ({str_inicio} al {str_fin}).xlsx"

        with pd.ExcelWriter(archivo_final, engine='openpyxl') as writer:
            df_datos.to_excel(writer, index=False, sheet_name='Consolidado')
            worksheet = writer.sheets['Consolidado']

            fill_encabezado = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            fuente_encabezado = Font(color="FFFFFF", bold=True)
            alineacion_centrada = Alignment(horizontal="center", vertical="center")

            for col_num in range(1, len(df_datos.columns) + 1):
                celda = worksheet.cell(row=1, column=col_num)
                celda.fill = fill_encabezado
                celda.font = fuente_encabezado
                celda.alignment = alineacion_centrada

            for idx, col_name in enumerate(df_datos.columns):
                ancho_calculado = max(
                    df_datos[col_name].astype(str).map(len).max(),
                    len(str(col_name))
                ) + 3
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(ancho_calculado, 40)

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

        logging.info("Renderizado exitoso. Archivo de salida disponible: %s", archivo_final)

    except Exception as e:
        logging.error("Fallo critico I/O durante la escritura del archivo Excel: %s", e)


if __name__ == "__main__":
    fecha_inicio, fecha_fin = solicitar_rango_fechas()
    token_jwt = obtener_token()

    if token_jwt:
        df_resultado, _ = procesar_cruce_datos(token_jwt, fecha_inicio, fecha_fin)
        exportar_a_excel_empresarial(df_resultado, fecha_inicio, fecha_fin)
    else:
        logging.critical("Secuencia interrumpida. No fue posible establecer conexion con el ERP.")